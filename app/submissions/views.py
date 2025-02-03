import json, os, re, copy
from io import BytesIO
import zipfile
from collections import Counter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.borders import Side
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas as reportlab_canvas
from reportlab.lib.pagesizes import letter as reportlab_letter
from reportlab.lib.colors import white as reportlab_white, black as reportlab_black, Color as reportlab_Color
from reportlab.platypus import Image as reportlab_Image
from reportlab.pdfbase import pdfmetrics as reportlab_pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont as reportlab_TTFont
from reportlab.lib import colors as reportlab_colors
from reportlab.platypus import SimpleDocTemplate as reportlab_SimpleDocTemplate, BaseDocTemplate as reportlab_BaseDocTemplate, PageTemplate as reportlab_PageTemplate, Frame as reportlab_Frame, Table as reportlab_Table, TableStyle as reportlab_TableStyle, Paragraph as reportlab_Paragraph, Spacer as reportlab_Spacer, PageBreak as reportlab_PageBreak, HRFlowable as reportlab_HRFlowable, Flowable as reportlab_Flowable
from reportlab.lib.styles import getSampleStyleSheet as reportlab_getSampleStyleSheet, ParagraphStyle as reportlab_ParagraphStyle
from django.db import models
from django.db.models import Min, Max, Sum, Count, Q, Prefetch
from django.db.models.functions import Coalesce, Lower
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.generic.edit import FormView, DeleteView
from django.views.decorators.http import require_http_methods
from django.core.mail import send_mail
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.conf import settings
from django.http import Http404, HttpResponse, StreamingHttpResponse
from django.utils.html import escape
from rest_framework import generics, viewsets, status
from rest_framework.views import APIView
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from users.models import User
from .models import STATE_CHOICES, Fair, CurrentFair, Languoid, Tribe, Submission, Category, Instructor, Student, Accessory, SubmissionAccessory
from .serializers import CategorySerializer, SubmissionSerializer, PosterSerializer, InstructorSerializer, StudentSerializer, SubmissionAccessorySerializer, SubmissionJsonSerializer
from .forms import SubmissionForm, SubmissionCommentsForm, InstructorForm, StudentForm, PosterForm
from users.utils import generate_registration_code
from datetime import timedelta


import requests
from django.http import JsonResponse
import logging
from django.db import transaction
from django.utils import timezone

logger = logging.getLogger(__name__)

def is_moderator(user):
    return user.groups.filter(name='moderator').exists()

def is_admin(user):
    return user.is_superuser or user.is_staff

def custom_500_view(request):
    return render(request, "500.html", status=500)

def contact_info(request):
    currentFair = CurrentFair.objects.first()

    template = 'contact_info.html'
    context = {
        'currentFair': currentFair.name,
    }
    return render(request, template, context)

@api_view(['GET'])
def submission_list(request):
    if not request.user.is_authenticated:
        return Response({'detail': 'Authentication credentials were not provided.'}, 
                       status=status.HTTP_401_UNAUTHORIZED)
    
    # Get the fair_id from query params, default to current fair
    fair_id = request.GET.get('fair_id')
    if fair_id:
        fair = get_object_or_404(Fair, pk=fair_id)
    else:
        current_fair = CurrentFair.objects.first()
        if not current_fair:
            return Response({'detail': 'No current fair set.'}, status=status.HTTP_404_NOT_FOUND)
        fair = current_fair.fair

    # Start with optimized query using select_related and prefetch_related
    submissions = (Submission.objects
        .select_related('user', 'category', 'fair')  # For foreign keys
        .prefetch_related(
            'students',
            'instructors',
            'languoids',
            'accessories'
        )
        .filter(fair=fair)
    )

    # Apply filters based on query parameters
    if user_id := request.GET.get('user_id'):
        submissions = submissions.filter(user_id=user_id)
    
    serializer = SubmissionSerializer(submissions, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def poster_list(request):
    if not request.user.is_authenticated:
        return Response({'detail': 'Authentication credentials were not provided.'}, status=status.HTTP_401_UNAUTHORIZED)
    submissions = Submission.objects.filter(poster=True)
    serializer = PosterSerializer(submissions, many=True)
    return Response(serializer.data)

# @api_view(['GET'])
# def submission_poster_list(request):
#     if not request.user.is_authenticated:
#         return Response({'detail': 'Authentication credentials were not provided.'}, status=status.HTTP_401_UNAUTHORIZED)
    
#     # Get the fair_id from query params, default to current fair
#     fair_id = request.GET.get('fair_id')
#     if fair_id:
#         fair = get_object_or_404(Fair, pk=fair_id)
#     else:
#         current_fair = CurrentFair.objects.first()
#         if not current_fair:
#             return Response({'detail': 'No current fair set.'}, status=status.HTTP_404_NOT_FOUND)
#         fair = current_fair.fair
    
#     # Filter submissions by fair and user if specified
#     submissions = Submission.objects.filter(fair=fair)
#     user_id = request.GET.get('user_id')
#     if user_id is not None:
#         user = get_object_or_404(User, pk=user_id)
#         submissions = submissions.filter(user=user)
    
#     serializer = SubmissionSerializer(submissions, many=True)
#     return Response(serializer.data)

@api_view(['GET'])
def submission_get(request, perf_pk):
    if not request.user.is_authenticated:
        return Response({'detail': 'Authentication credentials were not provided.'}, status=status.HTTP_401_UNAUTHORIZED)
        # Check if the user is in the "moderator" group
    submission = get_object_or_404(Submission, pk=perf_pk)
    if not request.user.groups.filter(name='moderator').exists():
        if not request.user == submission.user:
            return Response({'detail': 'You do not have permission to perform this action.'}, status=status.HTTP_403_FORBIDDEN)
    serializer = SubmissionSerializer(submission)
    return Response(serializer.data)

class CategoryUpdateView(LoginRequiredMixin, generics.UpdateAPIView):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    # permission_classes = [IsAuthenticated]  # Example: Allow any user to update tasks !!!!!!!!!!!!!!!

class InstructorAddView(APIView):
    def post(self, request):
        # Get the current fair
        currentFair = CurrentFair.objects.first()
        # Get the user id from the request data
        user_id = request.data.get('user_id')

        # Get the user from the request
        user = User.objects.get(id=user_id)

        # Create a new instructor with the data from the request
        instructor = Instructor(
            firstname=request.data.get('firstname'),
            lastname=request.data.get('lastname'),
            user=user,
            fair=currentFair.fair,
            modified_by=request.user.get_username(),
        )

        # Save the instructor
        instructor.save()

        # Serialize the instructor
        serializer = InstructorSerializer(instructor)

        # Return the serialized instructor
        return Response(serializer.data, status=status.HTTP_201_CREATED)

class InstructorUpdateView(APIView):
    def put(self, request, instr_pk):
        instructor = self.get_object(instr_pk)
        serializer = InstructorSerializer(instructor, data=request.data)
        if serializer.is_valid():
            serializer.validated_data['modified_by'] = request.user.get_username()
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get_object(self, instr_pk):
        try:
            return Instructor.objects.get(pk=instr_pk)
        except Instructor.DoesNotExist:
            raise Http404

class InstructorViewSet(LoginRequiredMixin, viewsets.ModelViewSet):
    serializer_class = InstructorSerializer

    def get_queryset(self):
        queryset = Instructor.objects.all()
        user_id = self.request.query_params.get('user_id', None)
        submission_id = self.request.query_params.get('submission_id', None)
        fair_id = self.request.query_params.get('fair_id', None)

        if fair_id is not None:
            queryset = queryset.filter(fair_id=fair_id)
        elif not submission_id:  # If no fair_id and no submission_id, use current fair
            current_fair = CurrentFair.objects.first()
            if current_fair:
                queryset = queryset.filter(fair=current_fair.fair)

        if user_id is not None:
            queryset = queryset.filter(user_id=user_id)

        if submission_id is not None:
            queryset = queryset.filter(submission_instructor__id=submission_id)
        
        return queryset

class StudentAddView(APIView):
    def post(self, request):
        # Get the current fair
        currentFair = CurrentFair.objects.first()
        # Get the user id from the request data
        user_id = request.data.get('user_id')

        # Get the user from the request
        user = User.objects.get(id=user_id)

        # Create a new student with the data from the request
        student = Student(
            firstname=request.data.get('firstname'),
            lastname=request.data.get('lastname'),
            grade=request.data.get('grade'),
            hometown=request.data.get('hometown'),
            state=request.data.get('state'),
            tshirt_size=request.data.get('tshirt_size'),
            user=user,
            fair=currentFair.fair,
            modified_by=request.user.get_username(),
        )

        # Save the student
        student.save()

        student.tribe.set(request.data.get('tribes'))

        # Serialize the student
        serializer = StudentSerializer(student)

        # Return the serialized student
        return Response(serializer.data, status=status.HTTP_201_CREATED)

class StudentUpdateView(APIView):
    def put(self, request, stud_pk):
        # Get the current fair
        currentFair = CurrentFair.objects.first()
        # Get the user id from the request data
        user_id = request.data.get('user_id')

        # Get the user from the request
        user = User.objects.get(id=user_id)

        # Get the student to update
        student = get_object_or_404(Student, pk=stud_pk)

        # Update the student with the data from the request
        student.firstname = request.data.get('firstname')
        student.lastname = request.data.get('lastname')
        student.grade = request.data.get('grade')
        student.hometown = request.data.get('hometown')
        student.state = request.data.get('state')
        student.tshirt_size = request.data.get('tshirt_size')
        # student.user = user # This line is commented out because the user should not be updated
        # student.fair = currentFair.fair # This line is commented out because the fair should not be updated
        student.modified_by = request.user.get_username()

        # Save the student
        student.save()

        student.tribe.set(request.data.get('tribes'))

        # Serialize the student
        serializer = StudentSerializer(student)

        # Return the serialized student
        return Response(serializer.data, status=status.HTTP_200_OK)


class StudentViewSet(LoginRequiredMixin, viewsets.ModelViewSet):
    serializer_class = StudentSerializer

    def get_queryset(self):
        queryset = Student.objects.all()
        user_id = self.request.query_params.get('user_id', None)
        submission_id = self.request.query_params.get('submission_id', None)
        fair_id = self.request.query_params.get('fair_id', None)

        if fair_id is not None:
            queryset = queryset.filter(fair_id=fair_id)
        elif not submission_id:  # If no fair_id and no submission_id, use current fair
            current_fair = CurrentFair.objects.first()
            if current_fair:
                queryset = queryset.filter(fair=current_fair.fair)

        if user_id is not None:
            queryset = queryset.filter(user_id=user_id)
        if submission_id is not None:
            queryset = queryset.filter(submission_student__id=submission_id)
        
        return queryset

class SubmissionUpdateView(LoginRequiredMixin, generics.UpdateAPIView):
    queryset = Submission.objects.all()
    serializer_class = SubmissionSerializer

class SubmissionAccessoryCreateView(LoginRequiredMixin, generics.CreateAPIView):
    queryset = SubmissionAccessory.objects.all()
    serializer_class = SubmissionAccessorySerializer

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def perform_create(self, serializer):
        serializer.save()

class SubmissionAccessoryUpdateView(LoginRequiredMixin, generics.UpdateAPIView):
    queryset = SubmissionAccessory.objects.all()
    serializer_class = SubmissionAccessorySerializer

    def get_object(self):
        submission_id = self.kwargs.get('perf_pk')
        accessory_id = self.kwargs.get('acc_pk')
        return get_object_or_404(SubmissionAccessory, submission=submission_id, accessory=accessory_id)

@login_required
def select_fair(request, pk=None):

    currentFair = CurrentFair.objects.first()

    if pk:
        get_fair_to_update = Fair.objects.get(id=pk)
        if get_fair_to_update:
            currentFair.fair = get_fair_to_update
            currentFair.name = get_fair_to_update.name
            currentFair.save()

    fairs = Fair.objects.values('id', 'name')
    fairs_ordered = fairs.order_by('-updated')

    recent_fairs = fairs.order_by('-updated') # this modification of fairs is currently redundant, but it would be needed if the ordering of fairs above needed to change
    if len(recent_fairs) > 3:
        recent_fairs = recent_fairs[:3]

    template = 'select_fair.html'
    context = {
        'currentFair': currentFair.name,
        'recent_fairs': recent_fairs,
        'fairs': fairs_ordered
    }
    return render(request, template, context)


class SubmissionAccessoryViewSet(LoginRequiredMixin, viewsets.ModelViewSet):
    queryset = SubmissionAccessory.objects.all()
    serializer_class = SubmissionAccessorySerializer

    def get_queryset(self):
        queryset = SubmissionAccessory.objects.all()
        submission_id = self.request.query_params.get('submission_id', None)
        accessory_id = self.request.query_params.get('accessory_id', None)

        if submission_id is not None:
            queryset = queryset.filter(submission__id=submission_id)
        if accessory_id is not None:
            queryset = queryset.filter(accessory__id=accessory_id)
        return queryset

@login_required
@user_passes_test(is_moderator)
def fair_list(request):
    fairs = Fair.objects.all()
    
    # Get the current fair (system setting)
    current_fair = CurrentFair.objects.first()
    
    # Get the active fair for the UI (either from URL parameter or default to current fair)
    active_fair_id = request.GET.get('fair_id')
    if active_fair_id:
        try:
            active_fair = Fair.objects.get(id=active_fair_id)
        except Fair.DoesNotExist:
            active_fair = current_fair.fair if current_fair else None
    else:
        active_fair = current_fair.fair if current_fair else None

    if active_fair:
        # Get related objects for the active fair
        languoids = active_fair.fair_languoids.all()
        tribes = active_fair.fair_tribes.all()
        categories = active_fair.fair_categories.all()
        accessories = active_fair.fair_accessories.all()
    else:
        languoids = tribes = categories = accessories = []

    context = {
        'fairs': fairs,
        'active_fair': active_fair,
        'current_fair': current_fair.fair if current_fair else None,  # For highlighting the system-wide current fair
        'languoids': languoids,
        'tribes': tribes,
        'categories': categories,
        'accessories': accessories,
    }
    return render(request, 'fair_list.html', context)

@login_required
@require_http_methods(["GET"])
def get_fair(request, pk):
    fair = get_object_or_404(Fair, pk=pk)
    current_fair = CurrentFair.objects.first()
    
    return JsonResponse({
        'id': fair.id,
        'name': fair.name,
        'notes': fair.notes,
        'registration_open': current_fair and current_fair.fair_id == fair.id,
        'is_current_fair': current_fair and current_fair.fair_id == fair.id,
        'languoids': [{
            'id': l.id,
            'name': l.name,
            'glottocode': l.glottocode,
            'level': l.level,
            'level_display': l.get_level_display()
        } for l in fair.fair_languoids.all()],
        'tribes': [{
            'id': t.id,
            'name': t.name
        } for t in fair.fair_tribes.all()],
        'categories': [{
            'id': c.id,
            'name': c.name,
            'material_submission': c.material_submission,
            'max_students': c.max_students
        } for c in fair.fair_categories.all()],
        'accessories': [{
            'id': a.id,
            'name': a.name
        } for a in fair.fair_accessories.all()]
    })

@login_required
@require_http_methods(["POST"])
def edit_fair(request, pk):
    fair = get_object_or_404(Fair, pk=pk)
    data = json.loads(request.body)
    
    if 'name' in data:
        fair.name = data['name']
    if 'notes' in data:
        fair.notes = data['notes']
    if 'registration_open' in data:
        fair.registration_open = data['registration_open']
    
    fair.modified_by = request.user.get_username()
    fair.save()
    
    return JsonResponse({'success': True})

def handle_related_item(model_class, fair_id, request, item_id=None):
    fair = get_object_or_404(Fair, pk=fair_id)
    
    if request.method == "POST":
        data = json.loads(request.body)
        item = model_class.objects.create(
            fair=fair,
            name=data['name'],
            modified_by=request.user.get_username()
        )
        return JsonResponse({'id': item.id, 'name': item.name})
        
    elif request.method == "PUT":
        item = get_object_or_404(model_class, pk=item_id, fair=fair)
        data = json.loads(request.body)
        item.name = data['name']
        item.modified_by = request.user.get_username()
        item.save()
        return JsonResponse({'id': item.id, 'name': item.name})
        
    elif request.method == "DELETE":
        item = get_object_or_404(model_class, pk=item_id, fair=fair)
        item.delete()
        return JsonResponse({'success': True})

@login_required
@require_http_methods(["GET"])
def check_delete_item(request, fair_id, type, item_id):
    fair = get_object_or_404(Fair, pk=fair_id)
    model_map = {
        'languoids': Languoid,
        'tribes': Tribe,
        'categories': Category,
        'accessories': Accessory,
    }
    
    model_class = model_map.get(type)
    if not model_class:
        return JsonResponse({'error': 'Invalid type'}, status=400)
        
    item = get_object_or_404(model_class, pk=item_id, fair=fair)
    associations = []
    
    # Check submissions using this item
    if type == 'languoids':
        submissions = Submission.objects.filter(languoids=item)
    elif type == 'tribes':
        submissions = Submission.objects.filter(students__tribe=item).distinct()
    elif type == 'categories':
        submissions = Submission.objects.filter(category=item)
    elif type == 'accessories':
        submissions = Submission.objects.filter(
            submissionaccessory__accessory=item,
            submissionaccessory__count__gt=0
        ).distinct()
    
    if submissions.exists():
        if type == 'accessories':
            associations.append({
                'type': 'Submissions',
                'items': [{
                    'id': sub.id,
                    'text': f"{sub.title} ({sub.organization or 'No organization'}, count: {sub.submissionaccessory_set.get(accessory=item).count})"
                } for sub in submissions]
            })
        else:
            associations.append({
                'type': 'Submissions',
                'items': [{
                    'id': sub.id,
                    'text': f"{sub.title} ({sub.organization or 'No organization'})"
                } for sub in submissions]
            })
    
    # Additional checks for tribes (students)
    if type == 'tribes':
        students = Student.objects.filter(tribe=item)
        if students.exists():
            associations.append({
                'type': 'Students',
                'items': [{'text': f"{student.firstname} {student.lastname}"} for student in students]
            })
    
    return JsonResponse({'associations': associations})

# Views for each related model type
@login_required
@require_http_methods(["POST", "PUT", "DELETE"])
def handle_languoid(request, fair_id, item_id=None):
    fair = get_object_or_404(Fair, pk=fair_id)
    
    if request.method == "POST":
        data = json.loads(request.body)
        languoid = Languoid.objects.create(
            fair=fair,
            name=data['name'],
            glottocode=data['glottocode'],
            level=data['level'],
            modified_by=request.user.get_username()
        )
        return JsonResponse({
            'id': languoid.id, 
            'name': languoid.name,
            'glottocode': languoid.glottocode,
            'level': languoid.level
        })
        
    elif request.method == "PUT":
        languoid = get_object_or_404(Languoid, pk=item_id, fair=fair)
        data = json.loads(request.body)
        languoid.name = data['name']
        languoid.glottocode = data['glottocode']
        languoid.level = data['level']
        languoid.modified_by = request.user.get_username()
        languoid.save()
        return JsonResponse({
            'id': languoid.id, 
            'name': languoid.name,
            'glottocode': languoid.glottocode,
            'level': languoid.level
        })
        
    elif request.method == "DELETE":
        languoid = get_object_or_404(Languoid, pk=item_id, fair=fair)
        languoid.delete()
        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Invalid request method'}, status=400)

@login_required
@require_http_methods(["POST", "PUT", "DELETE"])
def handle_tribe(request, fair_id, item_id=None):
    return handle_related_item(Tribe, fair_id, request, item_id)

@login_required
@require_http_methods(["POST", "PUT", "DELETE"])
def handle_category(request, fair_id, item_id=None):
    fair = get_object_or_404(Fair, pk=fair_id)
    
    if request.method == "POST":
        data = json.loads(request.body)
        category = Category.objects.create(
            fair=fair,
            name=data['name'],
            material_submission=data.get('material_submission', False),
            max_students=data.get('max_students'),  # This will handle None/null correctly
            modified_by=request.user.get_username()
        )
        return JsonResponse({
            'id': category.id, 
            'name': category.name,
            'material_submission': category.material_submission,
            'max_students': category.max_students
        })
        
    elif request.method == "PUT":
        category = get_object_or_404(Category, pk=item_id, fair=fair)
        data = json.loads(request.body)
        category.name = data['name']
        category.material_submission = data.get('material_submission', False)
        category.max_students = data.get('max_students')  # This will handle None/null correctly
        category.modified_by = request.user.get_username()
        category.save()
        return JsonResponse({
            'id': category.id, 
            'name': category.name,
            'material_submission': category.material_submission,
            'max_students': category.max_students
        })
        
    elif request.method == "DELETE":
        category = get_object_or_404(Category, pk=item_id, fair=fair)
        category.delete()
        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Invalid request method'}, status=400)

@login_required
@require_http_methods(["POST", "PUT", "DELETE"])
def handle_accessory(request, fair_id, item_id=None):
    return handle_related_item(Accessory, fair_id, request, item_id)

@login_required
@require_http_methods(["GET"])
def check_category_delete(request, fair_id, category_id):
    category = get_object_or_404(Category, id=category_id, fair_id=fair_id)
    associations = []
    
    # Check for submissions using this category
    submissions = Submission.objects.filter(category=category)
    if submissions.exists():
        submission_items = [f"{sub.title} (by {', '.join(str(student) for student in sub.students.all())})" 
                          for sub in submissions]
        associations.append({
            'type': 'Submissions',
            'items': submission_items
        })
    
    return JsonResponse({'associations': associations})

@login_required
@require_http_methods(["GET"])
def check_languoid_delete(request, fair_id, languoid_id):
    languoid = get_object_or_404(Languoid, id=languoid_id, fair_id=fair_id)
    associations = []
    
    # Check for submissions using this languoid
    submissions = Submission.objects.filter(languoids=languoid)
    if submissions.exists():
        submission_items = [f"{sub.title} (by {', '.join(str(student) for student in sub.students.all())})" 
                          for sub in submissions]
        associations.append({
            'type': 'Submissions',
            'items': submission_items
        })
    
    return JsonResponse({'associations': associations})

@login_required
@require_http_methods(["GET"])
def check_tribe_delete(request, fair_id, tribe_id):
    tribe = get_object_or_404(Tribe, id=tribe_id, fair_id=fair_id)
    associations = []
    
    # Check for students from this tribe
    students = Student.objects.filter(tribe=tribe)
    if students.exists():
        student_items = [str(student) for student in students]
        associations.append({
            'type': 'Students',
            'items': student_items
        })
        
    # Also check for submissions through these students
    submissions = Submission.objects.filter(students__tribe=tribe).distinct()
    if submissions.exists():
        submission_items = [f"{sub.title} (by {', '.join(str(student) for student in sub.students.all())})" 
                          for sub in submissions]
        associations.append({
            'type': 'Submissions',
            'items': submission_items
        })
    
    return JsonResponse({'associations': associations})

@login_required
@require_http_methods(["GET"])
def check_accessory_delete(request, fair_id, accessory_id):
    accessory = get_object_or_404(Accessory, id=accessory_id, fair_id=fair_id)
    
    # Get only the submissions that have this accessory with count > 0
    associated_submissions = Submission.objects.filter(
        submissionaccessory__accessory=accessory,
        submissionaccessory__count__gt=0  # Only count associations with count > 0
    ).distinct()
    
    if associated_submissions.exists():
        return JsonResponse({
            'associations': [{
                'type': 'Submissions',
                'items': [f"{sub.title} (count: {sub.submissionaccessory_set.get(accessory=accessory).count})" 
                         for sub in associated_submissions]
            }]
        })
    
    return JsonResponse({'associations': []})

@login_required
@user_passes_test(is_moderator)
def get_fair_data(request, fair_pk):
    try:
        fair = get_object_or_404(Fair, pk=fair_pk)
        data = {
            'id': fair.id,
            'name': fair.name,
            'notes': fair.notes or '',
            'registration_open': fair.registration_open,
            'languoids': [{
                'id': l.id,
                'name': l.name,
                'glottocode': l.glottocode,
                'level': l.level,
                'level_display': l.get_level_display()
            } for l in fair.fair_languoids.all()],
            'tribes': [{
                'id': t.id,
                'name': t.name
            } for t in fair.fair_tribes.all()],
            'categories': [{
                'id': c.id,
                'name': c.name,
                'material_submission': c.material_submission,
                'max_students': c.max_students
            } for c in fair.fair_categories.all()],
            'accessories': [{
                'id': a.id,
                'name': a.name
            } for a in fair.fair_accessories.all()]  # Changed 'l' to 'a' here
        }
        return JsonResponse(data)
    except Exception as e:
        logger.error(f"Error in get_fair_data: {str(e)}", exc_info=True)  # Added logging
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@user_passes_test(is_moderator)
@require_http_methods(["POST"])
def set_current_fair(request):
    try:
        data = json.loads(request.body)
        fair_id = data.get('fair_id')
        
        if not fair_id:
            return JsonResponse({'error': 'Fair ID is required'}, status=400)
            
        fair = get_object_or_404(Fair, id=fair_id)
        
        with transaction.atomic():
            # Set all fairs to registration_closed
            Fair.objects.all().update(registration_open=False)
            
            # Update or create current fair
            current_fair = CurrentFair.objects.first()
            if current_fair:
                current_fair.fair = fair
                current_fair.name = fair.name
                current_fair.save()
            else:
                CurrentFair.objects.create(fair=fair, name=fair.name)
            
        return JsonResponse({'success': True})
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@user_passes_test(is_moderator)
@require_http_methods(["POST"])
def add_fair(request):
    data = json.loads(request.body)
    name = data.get('name')
    notes = data.get('notes', '')
    template_fair_id = data.get('template_fair_id')
    
    if not name or not template_fair_id:
        return JsonResponse({'error': 'Name and template fair are required'}, status=400)
    
    try:
        with transaction.atomic():
            # Create new fair
            new_fair = Fair.objects.create(
                name=name,
                notes=notes,
                registration_open=False,
                modified_by=request.user.get_username()
            )
            
            # Get template fair
            template_fair = Fair.objects.get(id=template_fair_id)
            
            # Copy languoids
            for languoid in template_fair.fair_languoids.all():
                new_languoid = Languoid.objects.create(
                    fair=new_fair,
                    name=languoid.name,
                    glottocode=languoid.glottocode,
                    level=languoid.level,
                    modified_by=request.user.get_username()
                )
            
            # Copy tribes
            for tribe in template_fair.fair_tribes.all():
                new_tribe = Tribe.objects.create(
                    fair=new_fair,
                    name=tribe.name,
                    modified_by=request.user.get_username()
                )
            
            # Copy categories
            for category in template_fair.fair_categories.all():
                new_category = Category.objects.create(
                    fair=new_fair,
                    name=category.name,
                    material_submission=category.material_submission,
                    max_students=category.max_students,
                    modified_by=request.user.get_username()
                )
            
            # Copy accessories
            for accessory in template_fair.fair_accessories.all():
                new_accessory = Accessory.objects.create(
                    fair=new_fair,
                    name=accessory.name,
                    modified_by=request.user.get_username()
                )
            
            # Set all fairs to registration_closed
            Fair.objects.all().update(registration_open=False)
            
            # Update current fair
            current_fair = CurrentFair.objects.first()
            if current_fair:
                current_fair.fair = new_fair
                current_fair.name = new_fair.name
                current_fair.save()
            else:
                CurrentFair.objects.create(fair=new_fair, name=new_fair.name)
            
            return JsonResponse({
                'id': new_fair.id,
                'name': new_fair.name,
                'success': True
            })
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def home(request, fair_pk=None):

    currentUserEmail = request.user.get_username()
    currentUser = User.objects.get(email=currentUserEmail)
    
    # Check if the user is a moderator
    is_moderator = currentUser.groups.filter(name='moderator').exists()

    currentFair = CurrentFair.objects.first()


    if fair_pk is not None:
        fair = Fair.objects.get(pk=fair_pk)
    else:
        fair = currentFair.fair
    
    # If no fair exists, render a template informing the user
    if not fair:
        template = 'home.html'
        context = {
            'currentUser': currentUser,
            'moderator': is_moderator,
            'no_fair': True  # Add this flag to indicate no fair exists
        }
        return render(request, template, context)

    submissions = Submission.objects.prefetch_related("user", "category").filter(fair=fair)

    if not is_moderator:
        submissions = submissions.filter(user=currentUser)

    # Create the data first so we can inspect it
    data = {
        'gradeRanges': dict(Submission.GRADE_RANGES),
        'performanceTypes': dict(Submission.SUBMISSION_TYPE),
        'performanceStatus': dict(Submission.PERFORMANCE_STATUS),
        'categories': [
            {'id': cat.id, 'name': cat.name} 
            for cat in Category.objects.filter(fair=fair)
        ] if fair else []
    }

    # # Debug print
    # print("Data before JSON encoding:", data)

    # JSON encode the data
    json_data = {k: json.dumps(v) for k, v in data.items()}

    # # Debug print
    # print("Data after JSON encoding:", json_data)


    template = 'home.html'
    context = {
        'currentUser': currentUser,
        'moderator': is_moderator,
        'registrationOpen': fair.registration_open,
        'currentFair': currentFair.name,
        'submissions': submissions,
        'no_fair': fair is None,
        **json_data  # Add the JSON encoded data

    }
    return render(request, template, context)


@login_required
@user_passes_test(is_moderator)
def user_list(request):

    currentFair = CurrentFair.objects.first()

    # Check if the user is a moderator
    is_moderator = request.user.groups.filter(name='moderator').exists()

    # Get all fair names sorted alphabetically
    fair_names = Fair.objects.values_list('name', flat=True).order_by('name')
    fair_names_list = list(fair_names)

    # Get all users with their latest fair information
    all_users = User.objects.prefetch_related('groups').annotate(
        latest_fair=Coalesce(
            Max('submission_user__fair__name'),
            None
        )
    ).order_by(Lower('last_name'), Lower('first_name'))  # Add explicit ordering here

    # Add is_new flag and fair badge class for users
    six_months_ago = timezone.now() - timedelta(days=180)
    
    for user in all_users:
        user.is_new = user.date_joined >= six_months_ago
        
        # Add fair badge class based on position
        if user.latest_fair:
            try:
                position = fair_names_list[::-1].index(user.latest_fair)  # Reverse list for latest first
                if position == 0:
                    user.fair_badge_class = 'bg-success'
                elif position == 1:
                    user.fair_badge_class = 'bg-primary'
                elif position == 2:
                    user.fair_badge_class = 'bg-purple'
                else:
                    user.fair_badge_class = 'bg-secondary'
            except ValueError:
                user.fair_badge_class = 'bg-secondary'

    template = 'user_list.html'
    context = {
        'moderator': is_moderator,
        'currentFair': currentFair.name,
        'users': all_users,
        'state_choices': STATE_CHOICES,
    }

    if context['moderator']:
        context['registration_code'] = generate_registration_code()
    

    return render(request, template, context)


@login_required
@user_passes_test(is_moderator)
def user_detail(request, user_pk):

    currentUser = User.objects.get(pk=user_pk)

    currentFair = CurrentFair.objects.first()

    # Check if the user is a moderator
    is_moderator = request.user.groups.filter(name='moderator').exists()

    allSubmissions = Submission.objects.prefetch_related("user", "students").filter(fair=currentFair.fair).filter(user=currentUser.pk)

    submissions = allSubmissions.filter(poster=False)

    posters = allSubmissions.filter(poster=True)

    template = 'user_detail.html'
    context = {
        'currentUser': currentUser,
        'moderator': is_moderator,
        'currentFair': currentFair.name,
        'submissions': submissions,
        'posters': posters
    }
    return render(request, template, context)

@login_required
def submission_detail(request, perf_pk):

    currentFair = CurrentFair.objects.first()

    # Check if the user is a moderator
    is_moderator = request.user.groups.filter(name='moderator').exists()

    submission = get_object_or_404(
        Submission.objects.prefetch_related('instructors', 'students', 'accessories'),
        pk=perf_pk
    )

    # get the non material submission categories for the current fair
    non_material_submission_categories = list(Category.objects.filter(fair=currentFair.fair, material_submission=False).values_list('name', flat=True))
    # check if this submission is a non-material submission
    non_material_submission = submission.category.name in non_material_submission_categories

    # Optimized accessory queries
    accessories = Accessory.objects.filter(fair=currentFair.fair).prefetch_related(
        'submissionaccessory_set'
    )
    accessory_counts = {
        acc.id: acc.submissionaccessory_set.get(submission_id=perf_pk).count 
        for acc in accessories 
        if acc.submissionaccessory_set.filter(submission_id=perf_pk).exists()
    }

    # Update the count attribute of each accessory with the count from the submission_accessories
    for accessory in accessories:
        if accessory.id in accessory_counts:
            accessory.count = accessory_counts[accessory.id]

    # Get the "Other" languoid specific to this submission's fair
    other_languoid = Languoid.objects.get(name='Other', fair=submission.fair)

    submission_includes_other_languoid = submission.languoids.filter(pk=other_languoid.pk).exists()

    template = 'submission_detail.html'
    context = {
        'currentFair': currentFair.name,
        'moderator': is_moderator,
        'submission': submission,
        'non_material_submission': non_material_submission,
        'includes_other_languoid': submission_includes_other_languoid,
        'accessories': accessories,
    }
    return render(request, template, context)

class submission_add(LoginRequiredMixin, FormView):
    def handle_no_permission(self):
        return redirect('/accounts/login/')
    form_class = SubmissionForm
    template_name = "submission_add.html"

    def get_initial_category(self):
        return self.kwargs.get('category', None)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        current_fair = CurrentFair.objects.first()
        kwargs['current_fair'] = current_fair.fair
        kwargs['selected_category'] = self.get_initial_category()
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['add_type'] = 'Submission'

        if 'user_pk' in self.kwargs:
            user_pk = self.kwargs['user_pk']
            user = User.objects.get(pk=user_pk)
        else:
            user = self.request.user
        context['owning_user'] = user
        context['organization_preview'] = user.organization

        currentFair = CurrentFair.objects.first()
        context['currentFair'] = currentFair.name
        context['selected_category'] = self.get_initial_category()

        # Filter everything by current fair
        material_submission_categories = list(Category.objects.filter(
            fair=currentFair.fair, 
            material_submission=True
        ).values_list('name', flat=True))
        context['material_submission_categories'] = json.dumps(material_submission_categories)

        non_material_submission_categories = list(Category.objects.filter(
            fair=currentFair.fair, 
            material_submission=False
        ).values_list('name', flat=True))
        context['non_material_submission_categories'] = json.dumps(non_material_submission_categories)

        categories = Category.objects.filter(fair=currentFair.fair)
        categories_list = list(categories.values('name', 'max_students'))
        context['categories'] = json.dumps(categories_list)

        categories_dict = {category.name: category.max_students for category in categories}
        context['category_student_max'] = json.dumps(categories_dict)

        context['accessories'] = Accessory.objects.filter(fair=currentFair.fair)
        context['grades'] = Student.GRADES
        context['tribes'] = Tribe.objects.filter(fair=currentFair.fair)
        context['states'] = STATE_CHOICES
        context['tshirt_sizes'] = Student.TSHIRT_SIZES

        return context

    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)
    
    def form_valid(self, form):
        if form.is_valid():
            currentFair = CurrentFair.objects.first()
            self.object = form.save(commit=False)
            self.object.fair = currentFair.fair
            if 'user_pk' in self.kwargs:
                user_pk = self.kwargs['user_pk']  # Access the pk value
                user = User.objects.get(pk=user_pk)  # Get the user object
                self.object.user = user
            else:
                self.object.user = self.request.user
            if self.object.user.organization:
                self.object.organization = self.object.user.organization
            else:
                self.object.organization = ""
            self.object.modified_by = self.request.user.get_username()
            self.object.save()
            form.save_m2m()

            # Get the instructors from the form data
            instructors_json = self.request.POST.get('instructors')

            # Check if instructors_json exists
            if instructors_json:
                # Parse the instructors_json as a JSON array
                instructor_ids = json.loads(instructors_json)

                # For each instructor_id in the instructor_ids array...
                for instructor_id in instructor_ids:
                    # Get the Instructor object with the given ID
                    instructor = Instructor.objects.get(id=instructor_id)

                    # Add the instructor to the submission's instructors
                    self.object.instructors.add(instructor)

            # Get the students from the form data
            students_json = self.request.POST.get('students')

            # Check if students_json exists
            if students_json:
                # Parse the students_json as a JSON array
                student_ids = json.loads(students_json)

                # For each student_id in the student_ids array...
                for student_id in student_ids:
                    # Get the Student object with the given ID
                    student = Student.objects.get(id=student_id)

                    # Add the student to the submission's students
                    self.object.students.add(student)

            # Set submission_type based on the numner of students
            if self.object.students.count() > 1:
                self.object.submission_type = "group"
            else:
                self.object.submission_type = "individual"

            # get the student in the highest grade
            max_grade = self.object.students.aggregate(Max('grade'))
            max_grade_value = max_grade['grade__max']
            # look up the grade range based on the highest grade
            grade_range = Submission.GRADE_RANGES_DICT.get(max_grade_value)
            # set the grade range (based on the student in the highest grade)
            self.object.grade_range = grade_range

            # Save the submission object to update the instructors and students
            self.object.save()

            # Get the submission_accessory_counts from the form data
            submission_accessory_counts = self.request.POST.get('submission_accessory_counts')

            # Check if submission_accessory_counts exists
            if submission_accessory_counts:
                # Parse the submission_accessory_counts as a JSON object
                try:
                    accessory_counts = json.loads(submission_accessory_counts)
                except json.JSONDecodeError:
                    print(f"Invalid JSON: {submission_accessory_counts}")
                    accessory_counts = {}
                # For each key-value pair in the accessory_counts object...
                for accessory_id, count in accessory_counts.items():
                    # Create a SubmissionAccessory object
                    SubmissionAccessory.objects.create(
                        submission=self.object,
                        accessory_id=accessory_id,
                        count=count
                    )

            self.object.instructors_status = "completed"
            self.object.students_status = "completed"
            self.object.accessories_status = "completed"
            self.object.review_status = "completed"
            self.object.status = "submitted"
            self.object.save()

            # Determine the redirect URL based on the request path
            if 'submit-and-add' in self.request.POST:
                redirect_url = self.request.POST.get('submit-and-add')
                if redirect_url:
                    return redirect(redirect_url)
                
            return redirect("/submission/%s/" % self.object.pk)
    def form_invalid(self, form):
        print(form.errors)
        return super().form_invalid(form)

class submission_add_admin(UserPassesTestMixin, submission_add):
    def test_func(self):
        return self.request.user.groups.filter(name='moderator').exists()
    
@login_required
def submission_instructors(request, perf_pk):

    currentFair = CurrentFair.objects.first()

    submission = Submission.objects.get(pk=perf_pk)

    if submission.instructors_status == "pending":
        submission.instructors_status = "in_progress"
        submission.save()

    template = 'submission_instructors.html'
    context = {
        'currentFair': currentFair.name,
        'submission': submission
    }
    return render(request, template, context)

class instructor_add(LoginRequiredMixin, FormView):
    def handle_no_permission(self):
        return redirect('/no-permission')
    form_class = InstructorForm
    template_name = "instructor_add.html"
    def form_valid(self, form):
        if form.is_valid():
            currentFair = CurrentFair.objects.first()
            self.object = form.save(commit=False)
            self.object.fair = currentFair.fair
            if 'perf_pk' in self.kwargs:
                submission_id = self.kwargs['perf_pk']  # Access the pk value
                submission = Submission.objects.get(id=submission_id)
                self.object.user = submission.user
            else:
                self.object.user = self.request.user
            self.object.modified_by = self.request.user.get_username()
            self.object.save()
            # Determine the redirect URL based on the request path
            if 'submission/' in self.request.path:
                # return redirect(reverse('submission:submission_instructors_add', kwargs={'perf_pk': self.kwargs['perf_pk']}))
                return redirect("../")
            else:
                return redirect("../../")
    def form_invalid(self, form):
        print(form.errors)
        return super().form_invalid(form)

@login_required    
def instructor_edit(request, instr_pk, perf_pk=None):

    currentFair = CurrentFair.objects.first()

    instructor = get_object_or_404(Instructor, id=instr_pk)

    if request.method == "POST":
        form = InstructorForm(request.POST, instance=instructor)
        if form.is_valid():
            form.save(commit=False)
            form.modified_by = request.user.get_username()
            form.save()
            if perf_pk:
                return redirect("../../")
            else:
                return redirect("../../../")
    else:
        form = InstructorForm(instance=instructor)

    template = 'instructor_edit.html'
    context = {
        'currentFair': currentFair.name,
        'form': form
    }
    return render(request, template, context)

@login_required
def submission_students(request, perf_pk):

    currentFair = CurrentFair.objects.first()

    submission = Submission.objects.get(pk=perf_pk)

    if submission.students_status == "pending":
        submission.students_status = "in_progress"
        submission.save()

    template = 'submission_students.html'
    context = {
        'currentFair': currentFair.name,
        'submission': submission
    }
    return render(request, template, context)

class student_add(LoginRequiredMixin, FormView):
    def handle_no_permission(self):
        return redirect('/no-permission')
    form_class = StudentForm
    template_name = "student_add.html"
    def form_valid(self, form):
        if form.is_valid():
            currentFair = CurrentFair.objects.first()
            self.object = form.save(commit=False)
            self.object.fair = currentFair.fair
            if 'perf_pk' in self.kwargs:
                submission_id = self.kwargs['perf_pk']  # Access the pk value
                submission = Submission.objects.get(id=submission_id)
                self.object.user = submission.user
            else:
                self.object.user = self.request.user
            self.object.modified_by = self.request.user.get_username()
            self.object.save()
            # Determine the redirect URL based on the request path
            if 'submission/' in self.request.path:
                # return redirect(reverse('submission:submission_students_add', kwargs={'perf_pk': self.kwargs['perf_pk']}))
                return redirect("../")
            else:
                return redirect("../../")

            return redirect("../")
    def form_invalid(self, form):
        print(form.errors)
        return super().form_invalid(form)

@login_required
def student_edit(request, stud_pk, perf_pk=None):

    currentFair = CurrentFair.objects.first()

    student = get_object_or_404(Student, id=stud_pk)

    if request.method == "POST":
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save(commit=False)
            form.modified_by = request.user.get_username()
            form.save()
            return redirect("../../")
    else:
        form = StudentForm(instance=student)

    template = 'student_edit.html'
    context = {
        'currentFair': currentFair.name,
        'form': form
    }
    return render(request, template, context)

@login_required
def student_list(request):
    currentFair = CurrentFair.objects.first()

    # Get fair_id from query params, default to current fair if not provided
    fair_id = request.GET.get('fair_id')
    if fair_id:
        fair = get_object_or_404(Fair, id=fair_id)
    else:
        fair = currentFair.fair

    # Check if user is a moderator
    is_moderator = request.user.groups.filter(name='moderator').exists()

    # Optimize the prefetch for submissions to include category
    submission_prefetch = Prefetch(
        'submission_student',
        queryset=Submission.objects.select_related('category')
    )

    # Base queryset with optimized prefetching
    base_queryset = Student.objects.filter(fair=fair)\
        .select_related('user')\
        .prefetch_related(
            submission_prefetch,
            'tribe'
        )

    # Apply user filter if not moderator
    if not is_moderator:
        students = base_queryset.filter(user=request.user)
    else:
        students = base_queryset

    # Get all grades for filter dropdown
    grades = Student.GRADES

    # Get all categories for filter dropdown 
    categories = Category.objects.filter(fair=fair)

    template = 'student_list.html'
    context = {
        'currentFair': currentFair.name,
        'fair': fair,
        'students': students,
        'grades': grades,
        'categories': categories,
        'moderator': is_moderator
    }
    return render(request, template, context)


@login_required
def submission_accessories(request, perf_pk):

    currentFair = CurrentFair.objects.first()

    submission = Submission.objects.get(pk=perf_pk)

    if submission.accessories_status == "pending":
        submission.accessories_status = "in_progress"
        submission.save()

    # Fetch the SubmissionAccessory instances related to the current submission
    submission_accessories = SubmissionAccessory.objects.filter(submission=submission)

    # Create a dictionary mapping accessory IDs to counts
    accessory_counts = {pa.accessory.id: pa.count for pa in submission_accessories}

    # Fetch the accessories
    accessories = Accessory.objects.filter(fair=currentFair.fair)

    # Update the count attribute of each accessory with the count from the submission_accessories
    for accessory in accessories:
        if accessory.id in accessory_counts:
            accessory.count = accessory_counts[accessory.id]

    template = 'submission_accessories.html'
    context = {
        'currentFair': currentFair.name,
        'submission': submission,
        'accessories': accessories
    }
    return render(request, template, context)

@login_required
def submission_edit(request, perf_pk):

    # Check if the user is a moderator
    is_moderator = request.user.groups.filter(name='moderator').exists()

    currentFair = CurrentFair.objects.first()

    submission = get_object_or_404(Submission, id=perf_pk)

    owning_user = submission.user

    # if the submission is from a different fair, organization_preview is the organization already set
    if submission.fair != currentFair.fair:
        organization_preview = submission.organization
    else:
        organization_preview = owning_user.organization

    # If user is not a moderator and submission is from a different fair, redirect to detail
    if not is_moderator and submission.fair != currentFair.fair:
        return redirect('submission_detail', perf_pk=perf_pk)

    # Redirect if submission status is not 'in_progress'
    if submission.status != 'in_progress':
        return redirect('submission_detail', perf_pk=perf_pk)

    material_submission_categories = list(Category.objects.filter(
        fair=submission.fair, 
        material_submission=True
    ).values_list('name', flat=True))
    
    non_material_submission_categories = list(Category.objects.filter(
        fair=submission.fair, 
        material_submission=False
    ).values_list('name', flat=True))

    categories = Category.objects.filter(fair=submission.fair)
    categories_list = list(categories.values('name', 'max_students'))
    categories_dict = {category.name: category.max_students for category in categories}

    other_languoid = Languoid.objects.get(name='Other', fair=submission.fair)

    submission_includes_other_languoid = submission.languoids.filter(pk=other_languoid.pk).exists()

    # Get accessories with their counts for this specific submission
    accessories = []
    for acc in Accessory.objects.filter(fair=submission.fair):
        try:
            submission_acc = SubmissionAccessory.objects.get(
                submission=submission,
                accessory=acc
            )
            acc.count = submission_acc.count
        except SubmissionAccessory.DoesNotExist:
            acc.count = 0
        accessories.append(acc)

    grades = Student.GRADES

    tribes = Tribe.objects.filter(fair=currentFair.fair)

    states = STATE_CHOICES

    tshirt_sizes = Student.TSHIRT_SIZES

    if request.method == "POST":
        form = SubmissionForm(request.POST, instance=submission, current_fair=submission.fair)
        if form.is_valid():
            edited_submission = form.save(commit=False)
            if edited_submission.fair == currentFair.fair:
                # Only set organization if user has one
                if edited_submission.user.organization:
                    edited_submission.organization = edited_submission.user.organization
                else:
                    edited_submission.organization = ""
            edited_submission.modified_by = request.user.get_username()
            edited_submission.save()
            form.save_m2m()

            # Get the instructors from the form data
            instructors_json = request.POST.get('instructors')

            # Check if instructors_json exists
            if instructors_json:
                # Parse the instructors_json as a JSON array
                instructor_ids = json.loads(instructors_json)

                # Get the Instructor objects with the given IDs
                instructors = Instructor.objects.filter(id__in=instructor_ids)

                # Set the submission's instructors to match the provided list
                edited_submission.instructors.set(instructors)

            # Get the students from the form data
            students_json = request.POST.get('students')

            # Check if students_json exists
            if students_json:
                # Parse the students_json as a JSON array
                student_ids = json.loads(students_json)

                # Get the Student objects with the given IDs
                students = Student.objects.filter(id__in=student_ids)

                # Set the submission's students to match the provided list
                edited_submission.students.set(students)

            # if override_submission_type is checked, do nothing special with submission type
            if not edited_submission.override_submission_type:
            # otherwise: set submission_type based on the numner of students
                if edited_submission.students.count() > 1:
                    edited_submission.submission_type = "group"
                else:
                    edited_submission.submission_type = "individual"


            # get the student in the highest grade
            max_grade = edited_submission.students.aggregate(Max('grade'))
            max_grade_value = max_grade['grade__max']
            # look up the grade range based on the highest grade
            grade_range = Submission.GRADE_RANGES_DICT.get(max_grade_value)
            # set the grade range (based on the student in the highest grade)
            edited_submission.grade_range = grade_range

            # Save the submission object to update the instructors and students
            edited_submission.save()

            # Get the submission_accessory_counts from the form data
            submission_accessory_counts = request.POST.get('submission_accessory_counts')

            # Check if submission_accessory_counts exists
            if submission_accessory_counts:
                # Parse the submission_accessory_counts as a JSON object
                try:
                    accessory_counts = json.loads(submission_accessory_counts)
                except json.JSONDecodeError:
                    print(f"Invalid JSON: {submission_accessory_counts}")
                    accessory_counts = {}
                # For each key-value pair in the accessory_counts object...
                for accessory_id, count in accessory_counts.items():
                    # Create a SubmissionAccessory object
                    SubmissionAccessory.objects.update_or_create(
                        submission=edited_submission,
                        accessory_id=accessory_id,
                        defaults={'count': count}
                    )
            
            edited_submission.instructors_status = "completed"
            edited_submission.students_status = "completed"
            edited_submission.accessories_status = "completed"
            edited_submission.review_status = "completed"
            edited_submission.status = "submitted"
            edited_submission.save()

            # Check if this is a submit-and-add request
            if 'submit-and-add' in request.POST:
                redirect_url = request.POST.get('submit-and-add')
                if redirect_url:
                    return redirect(redirect_url)

            return redirect(".")
    else:
        form = SubmissionForm(instance=submission, current_fair=submission.fair)
    template = 'submission_add.html'
    context = {
        'moderator': is_moderator,
        'currentFair': currentFair.name,
        'submission': submission,
        'owning_user': owning_user,
        'organization_preview': organization_preview,
        'material_submission_categories': json.dumps(material_submission_categories),
        'non_material_submission_categories': json.dumps(non_material_submission_categories),
        'categories': json.dumps(categories_list),
        'category_student_max': json.dumps(categories_dict),
        'includes_other_languoid': submission_includes_other_languoid,
        'accessories': accessories,
        'grades': grades,
        'tribes': tribes,
        'states': states,
        'tshirt_sizes': tshirt_sizes,
        'form': form
    }
    return render(request, template, context)

@login_required
def submission_review(request, perf_pk):

    currentFair = CurrentFair.objects.first()

    submission = Submission.objects.prefetch_related("instructors", "students", "accessories").get(pk=perf_pk)

    if submission.review_status == "pending":
        submission.review_status = "in_progress"
        submission.save()

    # Optimized accessory queries
    accessories = Accessory.objects.filter(fair=currentFair.fair).prefetch_related(
        'submissionaccessory_set'
    )
    accessory_counts = {
        acc.id: acc.submissionaccessory_set.get(submission_id=perf_pk).count 
        for acc in accessories 
        if acc.submissionaccessory_set.filter(submission_id=perf_pk).exists()
    }

    # Update the count attribute of each accessory with the count from the submission_accessories
    for accessory in accessories:
        if accessory.id in accessory_counts:
            accessory.count = accessory_counts[accessory.id]
    
    other_languoid = Languoid.objects.get(name='Other')
    submission_includes_other_languoid = submission.languoids.filter(pk=other_languoid.pk).exists()

    if request.method == "POST":
        form = SubmissionCommentsForm(request.POST, instance=submission)
        if form.is_valid():
            submission_form = form.save(commit=False)
            submission_form.review_status = "completed"
            submission_form.modified_by = request.user.get_username()
            submission_form.save()
            return redirect("/")
    else:
        form = SubmissionCommentsForm(instance=submission)

    template = 'submission_review.html'
    context = {
        'currentFair': currentFair.name,
        'submission': submission,
        'includes_other_languoid': submission_includes_other_languoid,
        'form': form
    }
    return render(request, template, context)

class poster_add(LoginRequiredMixin, FormView):
    def handle_no_permission(self):
        return redirect('/no-permission')
    form_class = PosterForm
    template_name = "poster_add.html"
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if 'user_pk' in self.kwargs:
            user_pk = self.kwargs['user_pk']  # Access the pk value
            user = User.objects.get(pk=user_pk)  # Get the user object
        else:
            user = self.request.user
        context['owning_user'] = user
        return context
    def form_valid(self, form):
        if form.is_valid():
            if 'user_pk' in self.kwargs:
                user_pk = self.kwargs['user_pk']  # Access the pk value
                user = User.objects.get(pk=user_pk)  # Get the user object
            else:
                user = self.request.user
            currentFair = CurrentFair.objects.first()
            self.object = form.save(commit=False)
            self.object.fair = currentFair.fair
            self.object.poster = True
            self.object.user = user
            self.object.instructors_status = "completed"
            self.object.students_status = "completed"
            self.object.accessories_status = "completed"
            self.object.review_status = "completed"
            self.object.modified_by = self.request.user.get_username()
            self.object.save()
            # self.object.title = "Poster " + str(self.object.pk)  # Set the name here
            self.object.save()  # Save the object again to store the new name
            form.save_m2m()
            for submission_accessory_count in form.submission_accessory_counts:
                if submission_accessory_count['count'] > 0:
                    print(submission_accessory_count['accessory'])
                    submission_accessory = SubmissionAccessory(submission=self.object, accessory=submission_accessory_count['accessory'], count=submission_accessory_count['count'])
                    submission_accessory.save()
            return redirect("../../")
    def form_invalid(self, form):
        print(form.errors)
        return super().form_invalid(form)


class poster_add_admin(UserPassesTestMixin, poster_add):
    def test_func(self):
        return self.request.user.groups.filter(name='moderator').exists()

@login_required
def poster_detail(request, post_pk):

    currentFair = CurrentFair.objects.first()

    # Check if the user is a moderator
    is_moderator = request.user.groups.filter(name='moderator').exists()

    submission = Submission.objects.prefetch_related("instructors", "students").get(pk=post_pk)

    other_languoid = Languoid.objects.get(name='Other')

    submission_includes_other_languoid = submission.languoids.filter(pk=other_languoid.pk).exists()

    template = 'poster_detail.html'
    context = {
        'currentFair': currentFair.name,
        'moderator': is_moderator,
        'submission': submission,
        'includes_other_languoid': submission_includes_other_languoid
    }
    return render(request, template, context)


@login_required
def poster_edit(request, post_pk):
    currentFair = CurrentFair.objects.first()

    submission = Submission.objects.prefetch_related("instructors", "students", "accessories").get(pk=post_pk)

    # Redirect if submission status is not 'in_progress'
    if submission.status != 'in_progress':
        return redirect('poster_detail', post_pk=post_pk)

    owning_user = submission.user

    if request.method == "POST":
        form = PosterForm(request.POST, instance=submission)
        if form.is_valid():
            submission = form.save(commit=False)
            submission.instructors_status = "completed"
            submission.students_status = "completed"
            submission.accessories_status = "completed"
            submission.review_status = "completed"
            submission.modified_by = request.user.get_username()
            submission.save()
            return redirect("../")
    else:
        form = PosterForm(instance=submission)

    template = 'poster_edit.html'
    context = {
        'currentFair': currentFair.name,
        'submission': submission,
        'owning_user': owning_user,
        'form': form
    }
    return render(request, template, context)

@login_required
@user_passes_test(is_moderator)
def fair_detail(request):

    # Check if the user is a moderator
    is_moderator = request.user.groups.filter(name='moderator').exists()

    currentFair = CurrentFair.objects.first()

    # Determine which fair's data to show
    fair_id = request.GET.get('fair_id')
    if fair_id:
        fair = get_object_or_404(Fair, id=fair_id)
    else:
        fair = currentFair.fair

    # find the number of submissions that are approved
    submissions_approved = Submission.objects.filter(fair=fair).filter(status="approved")
    submissions_approved_count = submissions_approved.count()

    # find the number of submissions that are submitted
    submissions_submitted = Submission.objects.filter(fair=fair).filter(status="submitted")
    submissions_submitted_count = submissions_submitted.count()

    # find the number of submissions that are approved or submitted
    submissions_total_count = submissions_approved_count + submissions_submitted_count

    # find the number of students in submissions that are approved
    students_approved = Student.objects.filter(submission_student__in=submissions_approved).distinct().count()

    # find the number of students in submissions that are submitted
    students_submitted = Student.objects.filter(submission_student__in=submissions_submitted).distinct().count()

    # find the number of students in submissions that are approved or submitted
    submissions_both = submissions_approved | submissions_submitted
    students_total = Student.objects.filter(submission_student__in=submissions_both).distinct().count()

    # find the number of approved submissions in each category
    submissions_by_category = {}
    categories = Category.objects.filter(fair=fair)
    for category in categories:
        submissions_by_category[category.name] = {
            "approved": submissions_approved.filter(category=category).count(),
            "submitted": submissions_submitted.filter(category=category).count(),
            "all": submissions_approved.filter(category=category).count() + submissions_submitted.filter(category=category).count()
        }

    # find the number of submissions that are approved for each language, and store only the non-zero results
    submissions_by_language = {}
    languoids = Languoid.objects.filter(fair=fair)

    # Get the "Other" languoid
    other_languoid = languoids.filter(name='Other').first()

    # Process non-"Other" languoids as before
    for languoid in languoids.exclude(name='Other'):
        submissions_by_language[languoid.name] = {
            "approved": submissions_approved.filter(languoids=languoid).count(),
            "submitted": submissions_submitted.filter(languoids=languoid).count(),
            "all": submissions_approved.filter(languoids=languoid).count() + submissions_submitted.filter(languoids=languoid).count()
        }

    # Special processing for "Other" languoid submissions
    if other_languoid:
        # Get all submissions with "Other" languoid
        other_approved = submissions_approved.filter(languoids=other_languoid)
        other_submitted = submissions_submitted.filter(languoids=other_languoid)
        
        # Get unique other_languoid values
        other_languoid_values = set(
            list(other_approved.values_list('other_languoid', flat=True).distinct()) +
            list(other_submitted.values_list('other_languoid', flat=True).distinct())
        )
        
        # Create a temporary dictionary for other languages
        other_languages = {}
        
        # Process each unique other_languoid value
        for other_value in other_languoid_values:
            key = f"Other: {other_value or 'Blank'}"
            other_languages[key] = {
                "approved": other_approved.filter(other_languoid=other_value).count(),
                "submitted": other_submitted.filter(other_languoid=other_value).count(),
                "all": (other_approved.filter(other_languoid=other_value).count() + 
                       other_submitted.filter(other_languoid=other_value).count())
            }
        
        # Sort other languages and add them to main dictionary
        # First, get all keys except "Other: Blank"
        sorted_other_keys = sorted([k for k in other_languages.keys() if k != "Other: Blank"])
        # If "Other: Blank" exists, add it at the end
        if "Other: Blank" in other_languages:
            sorted_other_keys.append("Other: Blank")
        
        # Add sorted other languages to main dictionary
        for key in sorted_other_keys:
            submissions_by_language[key] = other_languages[key]

    # remove any languages with no submissions
    submissions_by_language = {k: v for k, v in submissions_by_language.items() if v['all'] != 0}

    # find the number of submissions that are approved for each grade range
    submissions_by_grade_range = {}
    grade_ranges = Submission.GRADE_RANGES
    for grade_range, grade_range_display_value in grade_ranges:
        submissions_by_grade_range[grade_range_display_value] = {
            "approved": submissions_approved.filter(grade_range=grade_range).count(),
            "submitted": submissions_submitted.filter(grade_range=grade_range).count(),
            "all": submissions_approved.filter(grade_range=grade_range).count() + submissions_submitted.filter(grade_range=grade_range).count()
        }
    
    # filter submissions_approved and submissions_submitted to only include submissions with a category that is not a material submission
    submissions_approved_non_material = submissions_approved.filter(category__material_submission=False)
    submissions_submitted_non_material = submissions_submitted.filter(category__material_submission=False)

    # find the total number of tshirts needed in each size for all the approved or submitted submissions
    tshirt_sizes_summary = {}
    tshirt_sizes = Student.TSHIRT_SIZES
    for tshirt_size, tshirt_size_display_value in tshirt_sizes:
        tshirt_sizes_summary[tshirt_size_display_value] = {
            "approved": submissions_approved_non_material.filter(students__tshirt_size=tshirt_size).values("students").distinct().count(),
            "submitted": submissions_submitted_non_material.filter(students__tshirt_size=tshirt_size).values("students").distinct().count(),
            "all": submissions_approved_non_material.filter(students__tshirt_size=tshirt_size).values("students").distinct().count() + submissions_submitted_non_material.filter(students__tshirt_size=tshirt_size).values("students").distinct().count()
        }

    # filter submissions_approved and submissions_submitted to only include submissions with a category that is a material submission
    submissions_approved_material = submissions_approved.filter(category__material_submission=True)
    submissions_submitted_material = submissions_submitted.filter(category__material_submission=True)


    ## find the number of students that are in an approved submission that is a material submission, but are not in a non-material submission
    # Get all students in approved submissions that are material submissions
    # students_in_approved_material_submissions = submissions_approved_material.values("students").distinct()
    students_in_approved_material_submissions = Student.objects.filter(submission_student__in=submissions_approved_material).distinct()

    # # print a list of students in approved material submissions
    # print('students_in_approved_material_submissions')
    # students_list = students_in_approved_material_submissions.values_list('id', 'firstname', 'lastname')
    # # Get the list of students sorted by id
    # sorted_students_list = sorted(students_list, key=lambda x: x[0])

    # # Pretty print the sorted list of students
    # for id, firstname, lastname in sorted_students_list:
    #     print(f"ID: {id}, First Name: {firstname}, Last Name: {lastname}")
    # print("length of sorted_students_list")
    # print(len(sorted_students_list))
    # print(students_in_approved_material_submissions.count())

    # Get all students in approved submissions that are non-material submissions
    # students_in_approved_non_material_submissions = submissions_approved_non_material.values("students").distinct()
    students_in_approved_non_material_submissions = Student.objects.filter(submission_student__in=submissions_approved_non_material).distinct()


    # students_list = students_in_approved_non_material_submissions.values_list('id', 'firstname', 'lastname')
    # # Get the list of students sorted by id
    # sorted_students_list = sorted(students_list, key=lambda x: x[0])

    # # Pretty print the sorted list of students
    # for id, firstname, lastname in sorted_students_list:
    #     print(f"ID: {id}, First Name: {firstname}, Last Name: {lastname}")
    # print("length of non material sorted_students_list")
    # print(len(sorted_students_list))

    # print(students_in_approved_non_material_submissions.count())

    # Find students that are in an approved submission that is a material submission, but are not in a non-material submission
    students_in_approved_material_not_in_non_material = students_in_approved_material_submissions.exclude(id__in=students_in_approved_non_material_submissions).distinct()

    # print('students_in_approved_material_not_in_non_material')
    # print(students_in_approved_material_not_in_non_material.count())
    # print("students in both")
    # print(students_in_approved_material_submissions.filter(id__in=students_in_approved_non_material_submissions).count())

    # # find the intersection of the two sets
    # print("intersection")
    # print(students_in_approved_material_submissions.filter(id__in=students_in_approved_non_material_submissions).count())

    # # find the difference of the two sets
    # print("difference")
    # print(students_in_approved_material_submissions.difference(students_in_approved_non_material_submissions).count())


    ## find the number of students that are in a submitted submission that is a material submission, but are not in a non-material submission
    # Get all students in submitted submissions that are material submissions
    students_in_submitted_material_submissions = submissions_submitted_material.values("students").distinct()

    # Get all students in submitted submissions that are non-material submissions
    students_in_submitted_non_material_submissions = submissions_submitted_non_material.values("students").distinct()

    # Find students that are in a submitted submission that is a material submission, but are not in a non-material submission
    students_in_submitted_material_not_in_non_material = students_in_submitted_material_submissions.exclude(id__in=students_in_submitted_non_material_submissions)

    # find number of students in material submissions for bag counts
    bag_count = {
            "approved": students_in_approved_material_not_in_non_material.count(),
            "submitted": students_in_submitted_material_not_in_non_material.count(),
            "all": students_in_approved_material_not_in_non_material.count() + students_in_submitted_material_not_in_non_material.count()
        }

    # # filter SubmissionAccessory instances to only include those related to submissions that are approved or submitted
    # submission_accessories_approved = SubmissionAccessory.objects.filter(submission__in=submissions_approved)
    # submission_accessories_submitted = SubmissionAccessory.objects.filter(submission__in=submissions_submitted)

    # # filter submission_accessories_approved and submission_accessories_submitted to only include those related to submissions with a category that is not a material submission
    # submission_accessories_approved_non_material = submission_accessories_approved.filter(submission__category__material_submission=False)
    # submission_accessories_submitted_non_material = submission_accessories_submitted.filter(submission__category__material_submission=False)

    # find the total number of accessories in each type of accessory for all the approved submissions
    accessories = Accessory.objects.filter(fair=fair).prefetch_related(
        'submissionaccessory_set'
    )
    accessories_summary = {}
    for accessory in accessories:
        approved_count = sum(sa.count for sa in accessory.submissionaccessory_set.filter(
            submission__status='approved'
        ))
        submitted_count = sum(sa.count for sa in accessory.submissionaccessory_set.filter(
            submission__status='submitted'
        ))
        accessories_summary[accessory.name] = {
            "approved": approved_count,
            "submitted": submitted_count,
            "all": approved_count + submitted_count
        }

    # Get programs (users with organizations) that have submissions
    programs_with_submissions = User.objects.exclude(organization='').filter(
        submission_user__in=submissions_both
    ).distinct()

    # Get program counts using the same base query
    programs_submitted_count = User.objects.exclude(organization='').filter(
        submission_user__in=submissions_submitted
    ).distinct().count()

    programs_approved_count = User.objects.exclude(organization='').filter(
        submission_user__in=submissions_approved
    ).distinct().count()

    programs_total_count = programs_with_submissions.count()

    # Initialize counters for prek-5 and 6-12 programs
    prek5_programs = {
        "approved": 0,
        "submitted": 0,
        "all": 0
    }
    grade612_programs = {
        "approved": 0,
        "submitted": 0, 
        "all": 0
    }

    # Count approved programs with prek-5 submissions
    prek5_programs["approved"] = programs_with_submissions.filter(
        submission_user__in=submissions_approved,
        submission_user__grade_range__in=['0_pk-2', '1_3-5']
    ).distinct().count()

    # Count submitted programs with prek-5 submissions
    prek5_programs["submitted"] = programs_with_submissions.filter(
        submission_user__in=submissions_submitted,
        submission_user__grade_range__in=['0_pk-2', '1_3-5']
    ).distinct().count()

    # Count total programs with prek-5 submissions
    prek5_programs["all"] = programs_with_submissions.filter(
        submission_user__in=submissions_both,
        submission_user__grade_range__in=['0_pk-2', '1_3-5']
    ).distinct().count()

    # Count approved programs with 6-12 submissions
    grade612_programs["approved"] = programs_with_submissions.filter(
        submission_user__in=submissions_approved,
        submission_user__grade_range__in=['1_6-8', '1_9-12']
    ).distinct().count()

    # Count submitted programs with 6-12 submissions  
    grade612_programs["submitted"] = programs_with_submissions.filter(
        submission_user__in=submissions_submitted,
        submission_user__grade_range__in=['1_6-8', '1_9-12']
    ).distinct().count()

    # Count total programs with 6-12 submissions
    grade612_programs["all"] = programs_with_submissions.filter(
        submission_user__in=submissions_both,
        submission_user__grade_range__in=['1_6-8', '1_9-12']
    ).distinct().count()


 # Get programs by language, including special handling for "Other"
    programs_by_language = {}
    
    # Process non-"Other" languoids
    for languoid in languoids.exclude(name='Other'):
        programs_by_language[languoid.name] = {
            "approved": User.objects.exclude(organization='').filter(
                submission_user__in=submissions_approved,
                submission_user__languoids=languoid
            ).distinct().count(),
            "submitted": User.objects.exclude(organization='').filter(
                submission_user__in=submissions_submitted,
                submission_user__languoids=languoid
            ).distinct().count(),
            "all": User.objects.exclude(organization='').filter(
                submission_user__in=submissions_both,
                submission_user__languoids=languoid
            ).distinct().count()
        }

    # Special processing for "Other" languoid submissions
    if other_languoid:
        # Get all submissions with "Other" languoid
        other_approved = submissions_approved.filter(languoids=other_languoid)
        other_submitted = submissions_submitted.filter(languoids=other_languoid)
        
        # Get unique other_languoid values
        other_languoid_values = set(
            list(other_approved.values_list('other_languoid', flat=True).distinct()) +
            list(other_submitted.values_list('other_languoid', flat=True).distinct())
        )
        
        # Create a temporary dictionary for other languages
        other_languages = {}
        
        # Process each unique other_languoid value
        for other_value in other_languoid_values:
            key = f"Other: {other_value or 'Blank'}"
            other_languages[key] = {
                "approved": User.objects.exclude(organization='').filter(
                    submission_user__in=other_approved,
                    submission_user__other_languoid=other_value
                ).distinct().count(),
                "submitted": User.objects.exclude(organization='').filter(
                    submission_user__in=other_submitted,
                    submission_user__other_languoid=other_value
                ).distinct().count(),
                "all": User.objects.exclude(organization='').filter(
                    submission_user__in=submissions_both,
                    submission_user__languoids=other_languoid,
                    submission_user__other_languoid=other_value
                ).distinct().count()
            }
        
        # Sort and add other languages same as before
        sorted_other_keys = sorted([k for k in other_languages.keys() if k != "Other: Blank"])
        if "Other: Blank" in other_languages:
            sorted_other_keys.append("Other: Blank")
        
        for key in sorted_other_keys:
            programs_by_language[key] = other_languages[key]

    # Remove languages with no programs
    programs_by_language = {k: v for k, v in programs_by_language.items() if v['all'] != 0}


    # Get students by language, including special handling for "Other"
    students_by_language = {}
    
    # Process non-"Other" languoids
    for languoid in languoids.exclude(name='Other'):
        students_by_language[languoid.name] = {
            "approved": Student.objects.filter(
                submission_student__in=submissions_approved,
                submission_student__languoids=languoid
            ).distinct().count(),
            "submitted": Student.objects.filter(
                submission_student__in=submissions_submitted,
                submission_student__languoids=languoid
            ).distinct().count(),
            "all": Student.objects.filter(
                submission_student__in=submissions_both,
                submission_student__languoids=languoid
            ).distinct().count()
        }

    # Special processing for "Other" languoid submissions
    if other_languoid:
        # Get all submissions with "Other" languoid
        other_approved = submissions_approved.filter(languoids=other_languoid)
        other_submitted = submissions_submitted.filter(languoids=other_languoid)
        
        # Get unique other_languoid values
        other_languoid_values = set(
            list(other_approved.values_list('other_languoid', flat=True).distinct()) +
            list(other_submitted.values_list('other_languoid', flat=True).distinct())
        )
        
        # Create a temporary dictionary for other languages
        other_languages = {}
        
        # Process each unique other_languoid value
        for other_value in other_languoid_values:
            key = f"Other: {other_value or 'Blank'}"
            other_languages[key] = {
                "approved": Student.objects.filter(
                    submission_student__in=other_approved,
                    submission_student__other_languoid=other_value
                ).distinct().count(),
                "submitted": Student.objects.filter(
                    submission_student__in=other_submitted,
                    submission_student__other_languoid=other_value
                ).distinct().count(),
                "all": Student.objects.filter(
                    submission_student__in=submissions_both,
                    submission_student__languoids=other_languoid,
                    submission_student__other_languoid=other_value
                ).distinct().count()
            }
        
        # Sort and add other languages
        sorted_other_keys = sorted([k for k in other_languages.keys() if k != "Other: Blank"])
        if "Other: Blank" in other_languages:
            sorted_other_keys.append("Other: Blank")
        
        for key in sorted_other_keys:
            students_by_language[key] = other_languages[key]

    # Remove languages with no students
    students_by_language = {k: v for k, v in students_by_language.items() if v['all'] != 0}


    # Get all languoids for the fair
    languoids = Languoid.objects.filter(fair=fair)
    
    # Get the "Other" languoid
    other_languoid = languoids.filter(name='Other').first()

    # Initialize counters for prek-5 and 6-12 languages
    prek5_languages = {
        "approved": 0,
        "submitted": 0,
        "all": 0
    }
    grade612_languages = {
        "approved": 0,
        "submitted": 0, 
        "all": 0
    }

    # Helper function to count languages
    def count_languages(submissions_queryset, grade_ranges):
        # Count non-"Other" languoids
        language_count = languoids.exclude(name='Other').filter(
            submission_languoids__in=submissions_queryset,
            submission_languoids__grade_range__in=grade_ranges
        ).distinct().count()

        # If there's an "Other" languoid, count distinct other_languoid values
        if other_languoid:
            other_submissions = submissions_queryset.filter(
                languoids=other_languoid,
                grade_range__in=grade_ranges
            )
            # Get unique other_languoid values (including None/blank)
            other_values = set(other_submissions.values_list('other_languoid', flat=True).distinct())
            # Add the count of unique other_languoid values
            language_count += len(other_values)

        return language_count

    # Count languages for PreK-5 submissions
    prek5_languages["approved"] = count_languages(
        submissions_approved,
        ['0_pk-2', '1_3-5']
    )

    prek5_languages["submitted"] = count_languages(
        submissions_submitted,
        ['0_pk-2', '1_3-5']
    )

    prek5_languages["all"] = count_languages(
        submissions_both,
        ['0_pk-2', '1_3-5']
    )

    # Count languages for 6-12 submissions
    grade612_languages["approved"] = count_languages(
        submissions_approved,
        ['1_6-8', '1_9-12']
    )

    grade612_languages["submitted"] = count_languages(
        submissions_submitted,
        ['1_6-8', '1_9-12']
    )

    grade612_languages["all"] = count_languages(
        submissions_both,
        ['1_6-8', '1_9-12']
    )


    # Count all languages (across all grade ranges)
    languages_submitted_count = count_languages(
        submissions_submitted,
        ['0_pk-2', '1_3-5', '1_6-8', '1_9-12']
    )

    languages_approved_count = count_languages(
        submissions_approved,
        ['0_pk-2', '1_3-5', '1_6-8', '1_9-12']
    )

    languages_total_count = count_languages(
        submissions_both,
        ['0_pk-2', '1_3-5', '1_6-8', '1_9-12']
    )

    template = 'fair_detail.html'
    context = {
        'currentFair': currentFair.name,
        'fair': fair,
        'moderator': is_moderator,
        'submissions_submitted_count': submissions_submitted_count,
        'submissions_approved_count': submissions_approved_count,
        'submissions_total_count': submissions_total_count,
        'students_approved': students_approved,
        'students_submitted': students_submitted,
        'students_total': students_total,
        'submissions_by_category': submissions_by_category,
        'submissions_by_language': submissions_by_language,
        'submissions_by_grade_range': submissions_by_grade_range,
        'tshirt_sizes_summary': tshirt_sizes_summary,
        'bag_count': bag_count,
        'accessories_summary': accessories_summary,
        'prek5_programs': prek5_programs,
        'grade612_programs': grade612_programs,
        'prek5_languages': prek5_languages,
        'programs_by_language': programs_by_language,
        'grade612_languages': grade612_languages,
        'programs_submitted_count': programs_submitted_count,
        'programs_approved_count': programs_approved_count,
        'programs_total_count': programs_total_count,
        'students_by_language': students_by_language,
        'languages_submitted_count': languages_submitted_count,
        'languages_approved_count': languages_approved_count,
        'languages_total_count': languages_total_count,
    }
    return render(request, template, context)

# and API view that returns JSON for all the submissions for the fair given by the fair_pk, with all the metadata for each submission. This is sent to the browser as a download when the user clicks the "Download All Submission data" button on the fair detail page.
class FairDownloadView(APIView):
    def get(self, request, fair_pk):
        try:
            fair = Fair.objects.get(pk=fair_pk)
            submissions = Submission.objects.filter(fair=fair)

            ## Make json document
            serializer = SubmissionJsonSerializer(submissions, many=True)

            # Convert the serialized data to JSON and save it to a file
            data = json.dumps(serializer.data)
            json_file_name = f'fair-{fair.name}-data.json'
            json_file = default_storage.save(json_file_name, ContentFile(data))


            ## Adjust width of columns
            def adjust_width(sheet):
                limit_to_max_length = 40
                # Iterate over the columns
                for column in range(1, 8):
                    max_length = 0
                    column = get_column_letter(column)
                    for cell in sheet[column]:
                        try: 
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                                if max_length > limit_to_max_length:
                                    break
                        except:
                            pass
                        if max_length > limit_to_max_length:
                            break
                    adjusted_width = (max_length + 1)
                    if adjusted_width > limit_to_max_length:
                        adjusted_width = limit_to_max_length
                    sheet.column_dimensions[column].width = adjusted_width

            ## Make xlsx for submissions (non material)

            # filter submissions to only include those that are approved
            submissions = submissions.filter(status__in=["approved"])

            non_material_submission_categories = list(Category.objects.filter(fair=fair, material_submission=False).values_list('name', flat=True))
            categories = list(Category.objects.filter(fair=fair).values_list('name', flat=True))


            submission_workbook = Workbook()

            # # Define your default font
            # default_font = Font(name='Arial')

            # # Apply the default font to all cells in the workbook
            # for sheet in submission_workbook:
            #     for row in sheet.iter_rows():
            #         for cell in row:
            #             cell.font = default_font



            for category in categories:
                # sanitize the category name to remove any characters that are not allowed in a sheet name
                category_name = re.sub(r'[\\/*?[\]:]', '_', category)
                # create a new sheet
                submission_workbook.create_sheet(title=category_name)
                # make the sheet active
                submission_workbook.active = submission_workbook[category_name]
                sheet = submission_workbook.active
                # merge the first 7 columns of the first row
                sheet.merge_cells('A1:G1')
                # set the value of the merged cells to the category name
                sheet['A1'] = category
                # set the font size of the merged cells to 20 and bold
                sheet['A1'].font = Font(size=20, bold=True)
                # set the height of the first row to 30
                sheet.row_dimensions[1].height = 30
                # set the background color of the first 7 columns of the second row to blue, and the text color to white
                for cell in sheet['A2:G2']:
                    for c in cell:
                        c.fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
                        c.font = Font(color="FFFFFF")
                # list of column headers
                headers = ["Title", "Program/School", "Presenting group", "Language", "Grade Range", "Submission type", "Student count"]
                # set the values of the second row to the column headers
                for i, header in enumerate(headers):
                    sheet.cell(row=2, column=i+1, value=header)
                ## add data to the sheet, starting at the third row. the data are non material submissions with the current category
                # Iterate over the submissions in the category
                submissions_in_category = submissions.filter(category__name=category)
                for submission in submissions_in_category:
                    # Create a list for the current row
                    row = [
                        submission.title,
                        submission.organization,
                        submission.group,
                        ", ".join([f"Other: {submission.other_languoid if submission.other_languoid else 'Blank'}" if languoid.name == 'Other' else languoid.name for languoid in submission.languoids.all()]),
                        submission.get_grade_range_display(),
                        submission.get_submission_type_display(),
                        submission.students.count()
                    ]
                    # Add the row to the data list
                    sheet.append(row)

                adjust_width(sheet)

            # remove the default sheet
            submission_workbook.remove(submission_workbook['Sheet'])

            # Create a BytesIO object and save the workbook to it
            xlsx_file_io = BytesIO()
            submission_workbook.save(xlsx_file_io)

            # Go back to the start of the BytesIO object
            xlsx_file_io.seek(0)

            # Save the BytesIO object to a file in default storage
            submission_xlsx_file_name = f'Fair{fair.name}-Submission counts.xlsx'
            submission_xlsx_file = default_storage.save(submission_xlsx_file_name, ContentFile(xlsx_file_io.read()))


            # get all students in submissions for the current fair that are approved
            students = Student.objects.filter(submission_student__fair=fair).filter(submission_student__status="approved").distinct()

            student_workbook = Workbook()

            student_sorting_tabs = ["Sorted by name", "Sorted by age"]

            for tab in student_sorting_tabs:

                student_workbook.create_sheet(title=tab)
                student_workbook.active = student_workbook[tab]
                sheet = student_workbook.active

                # set the background color of the first 7 columns of the first row to blue, and the text color to white
                for cell in sheet['A1:G1']:
                    for c in cell:
                        c.fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
                        c.font = Font(color="FFFFFF")
                # list of column headers
                headers = ["First name", "Last name", "Tribe", "Hometown", "State", "Grade", "Program/School"]
                # set the values of the second row to the column headers
                for i, header in enumerate(headers):
                    sheet.cell(row=1, column=i+1, value=header)
                ## add data to the sheet, starting at the second row. the data are students in submissions for the current fair that are approved
                # Iterate over the students
                if tab == "Sorted by name":
                    students = students.order_by('lastname', 'firstname', 'grade')
                elif tab == "Sorted by age":
                    students = students.order_by('grade', 'lastname', 'firstname')
                for student in students:
                    # Create a list for the current row
                    row = [
                        student.firstname,
                        student.lastname,
                        ", ".join([tribe.name for tribe in student.tribe.all()]),
                        student.hometown,
                        student.state,
                        student.get_grade_display(),
                        student.user.organization
                    ]
                    # Add the row to the data list
                    sheet.append(row)

                adjust_width(sheet)

            # remove the default sheet
            student_workbook.remove(student_workbook['Sheet'])

            # Create a BytesIO object and save the workbook to it
            xlsx_file_io = BytesIO()
            student_workbook.save(xlsx_file_io)

            # Go back to the start of the BytesIO object
            xlsx_file_io.seek(0)

            # Save the BytesIO object to a file in default storage
            student_xlsx_file_name = f'Fair{fair.name}-Student details.xlsx'
            student_xlsx_file = default_storage.save(student_xlsx_file_name, ContentFile(xlsx_file_io.read()))

            # get all users that have approved submissions in the current fair and are not moderators
            users = User.objects.filter(
                submission_user__fair=fair,  # submission is in current fair
                submission_user__status='approved'  # submission is approved
            ).exclude(
                groups__name='moderator'
            ).distinct().order_by('organization', 'last_name', 'first_name')


            group_contact_workbook = Workbook()

            group_contact_workbook.create_sheet(title="GroupContactDetails")
            group_contact_workbook.active = group_contact_workbook["GroupContactDetails"]
            sheet = group_contact_workbook.active

            # set the background color of the first 7 columns of the first row to blue, and the text color to white
            for cell in sheet['A1:G1']:
                for c in cell:
                    c.fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
                    c.font = Font(color="FFFFFF")
            # list of column headers
            headers = ["Program/School", "Contact name", "Address", "City, State ZIP", "Phone number", "Fax number", "Email"]
            # set the values of the second row to the column headers
            for i, header in enumerate(headers):
                sheet.cell(row=1, column=i+1, value=header)
            for user in users:
                # Create a list for the current row
                row = [
                    user.organization or '',
                    " ".join(filter(None, [user.first_name, user.last_name])),
                    user.address or '',
                    ", ".join(filter(None, [user.city, " ".join(filter(None, [user.state, user.zip]))])),
                    ", ".join(filter(None, [user.phone, user.alt_phone])),
                    user.fax or '',
                    user.email or ''
                ]
                # Add the row to the data list
                sheet.append(row)

            adjust_width(sheet)

            # remove the default sheet
            group_contact_workbook.remove(group_contact_workbook['Sheet'])

            # Create a BytesIO object and save the workbook to it
            xlsx_file_io = BytesIO()
            group_contact_workbook.save(xlsx_file_io)

            # Go back to the start of the BytesIO object
            xlsx_file_io.seek(0)

            # Save the BytesIO object to a file in default storage
            group_contact_xlsx_file_name = f'Fair{fair.name}-Program contact details.xlsx'
            group_contact_xlsx_file = default_storage.save(group_contact_xlsx_file_name, ContentFile(xlsx_file_io.read()))



            # get a list of all the accessories for the fair
            accessories = Accessory.objects.filter(fair=fair)

            # Get all submissions
            submissions = Submission.objects.filter(fair=fair, status__in=["approved"]).values('id', 'title', 'organization', 'grade_range', 'category__name')
            
            # Convert GRADE_RANGES to a dictionary
            grade_ranges_dict = dict(Submission.GRADE_RANGES)

            # Replace the grade range with the display version
            for submission in submissions:
                submission['grade_range'] = grade_ranges_dict.get(submission['grade_range'])

            # Convert QuerySet to list of dictionaries
            submissions = list(submissions)

            # Loop over each submission
            for submission in submissions:
                # Loop over each accessory
                for accessory in accessories:
                    # Get the count of the current accessory for the current submission
                    count = SubmissionAccessory.objects.filter(submission_id=submission['id'], accessory=accessory).aggregate(count=Sum('count'))['count']

                    # Add the count to the submission dictionary
                    submission[f'accessory_{accessory.id}'] = count if count else ""

            accessory_workbook = Workbook()

            accessory_workbook.create_sheet(title="AccessoryCounts")
            accessory_workbook.active = accessory_workbook["AccessoryCounts"]
            sheet = accessory_workbook.active

            # list of column headers
            headers = ["Submission title", "Program/School", "Grade range", "Category"]
            # add the accessory names to the headers list
            headers.extend([accessory.name for accessory in accessories])

            # set the background color of the first row to blue, for the number of columnes equal to the length of the headers list
            for cell in sheet['A1':get_column_letter(len(headers))+'1']:
                for c in cell:
                    c.fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
                    c.font = Font(color="FFFFFF")

            # set the values of the second row to the column headers
            for i, header in enumerate(headers):
                sheet.cell(row=1, column=i+1, value=header)

            for submission in submissions:
                # Get all accessory counts for this submission
                accessory_counts = [submission[f'accessory_{accessory.id}'] for accessory in accessories]
                
                # Only add row if any accessory count is greater than 0
                if any(count and count > 0 for count in accessory_counts):
                    # Create a list for the current row
                    row = [
                        submission['title'],
                        submission['organization'],
                        submission['grade_range'],
                        submission['category__name']
                    ]
                    # add the accessory counts to the row
                    row.extend(accessory_counts)
                    
                    # Add the row to the data list
                    sheet.append(row)

            adjust_width(sheet)

            # remove the default sheet
            accessory_workbook.remove(accessory_workbook['Sheet'])

            # Create a BytesIO object and save the workbook to it
            xlsx_file_io = BytesIO()
            accessory_workbook.save(xlsx_file_io)

            # Go back to the start of the BytesIO object
            xlsx_file_io.seek(0)

            # Save the BytesIO object to a file in default storage
            accessory_xlsx_file_name = f'Fair{fair.name}-Accessory counts.xlsx'
            accessory_xlsx_file = default_storage.save(accessory_xlsx_file_name, ContentFile(xlsx_file_io.read()))


            # get all users that have approved submissions in the current fair and are not moderators
            users = User.objects.filter(
                submission_user__fair=fair,  # submission is in current fair
                submission_user__status='approved'  # submission is approved
            ).exclude(
                groups__name='moderator'
            ).distinct().order_by('organization', 'last_name', 'first_name')

            # Convert GRADES to a dictionary
            grades_dict = dict(Student.GRADES)
            grades_keys = list(grades_dict.keys())
            
            # make a list of dictionaries for each user
            users_list = []
            for user in users:
                # get the submissions for this user
                submissions = Submission.objects.filter(user=user, fair=fair)

                # get the students that are in the submissions
                students = Student.objects.filter(submission_student__in=submissions).distinct()

                # get the grades of the students
                grades = sorted(list(set(students.values_list('grade', flat=True))))

                ### deprecated code, using grade_range instead
                # # set day to "Day 1" if any of the grades are in grades_dict[0:6]
                # # set day to "Day 2" if any of the grades are in grades_dict[7:13]
                # # set day to "Day 1 + Day 2" if any of the grades are in grades_dict
                # day1 = "Day 1" if any(grade in grades_keys[0:7] for grade in grades) else ""
                # day2 = "Day 2" if any(grade in grades_keys[7:14] for grade in grades) else ""
                # day = " + ".join(filter(None, [day1, day2]))


                # Get all grade ranges from the user's submissions for this fair
                submission_grade_ranges = set(submissions.values_list('grade_range', flat=True))

                # Check for Day 1 grade ranges (PreK-2nd and 3rd-5th)
                day1_ranges = {'0_pk-2', '1_3-5'}
                day1 = "Day 1" if any(grade_range in day1_ranges for grade_range in submission_grade_ranges) else ""

                # Check for Day 2 grade ranges (6th-8th and 9th-12th)
                day2_ranges = {'1_6-8', '1_9-12'}
                day2 = "Day 2" if any(grade_range in day2_ranges for grade_range in submission_grade_ranges) else ""

                day = " + ".join(filter(None, [day1, day2]))




                # for each value in grades, replace with the display value from grades_dict
                grades = [grades_dict[grade] for grade in grades]

                # create a dictionary for the user
                user_dict = {
                    'name': f"{user.first_name} {user.last_name}",
                    'organization': user.organization,
                    'city': user.city,
                    'state': user.state,
                    'email': ', '.join(filter(None, [user.email, user.alt_email])),
                    'grades': ", ".join([str(grade) for grade in grades]),
                    'day': day,
                    'day1': day1,
                    'day2': day2
                }
                # add a copy of the dictionary to the list of users
                users_list.append(copy.deepcopy(user_dict))

            program_labels_workbook = Workbook()

            program_labels_workbook.create_sheet(title="Program Labels")
            program_labels_workbook.active = program_labels_workbook["Program Labels"]
            sheet = program_labels_workbook.active

            # list of column headers
            headers = ["Program/School", "Contact", "Day", "Grades", "Address"]

            # set the background color of the first row to blue, for the number of columnes equal to the length of the headers list
            for cell in sheet['A1':'E1']:
                for c in cell:
                    c.fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
                    c.font = Font(color="FFFFFF")

            # set the values of the second row to the column headers
            for i, header in enumerate(headers):
                sheet.cell(row=1, column=i+1, value=header)
            for user_dict in users_list:
                # Create contact string with name and email
                contact = f"{user_dict['name']} ({user_dict['email']})" if user_dict['name'].strip() else user_dict['email']
                
                # Create a list for the current row
                row = [
                    user_dict['organization'],
                    contact,
                    user_dict['day'],
                    user_dict['grades'],
                    f"{user_dict['city']}, {user_dict['state']}"
                ]
                
                # Add the row to the data list
                sheet.append(row)

            adjust_width(sheet)


            program_labels_workbook.create_sheet(title="Sign In Day 1")
            program_labels_workbook.active = program_labels_workbook["Sign In Day 1"]
            sheet = program_labels_workbook.active

            # list of column headers
            headers = ["Program/School", "Contact", "Location", "Registration Packet", "T-shirts/Bags", "Posters"]

            # set the background color of the first row to blue, for the number of columnes equal to the length of the headers list
            for cell in sheet['A1':'F1']:
                for c in cell:
                    c.font = Font(bold=True)

            # set the values of the second row to the column headers
            for i, header in enumerate(headers):
                sheet.cell(row=1, column=i+1, value=header)

            # filter the users_list to only include users that are in submissions that are on day 1
            users_list_day1 = [user_dict for user_dict in users_list if user_dict['day1'] == "Day 1"]
            for user_dict in users_list_day1:
                # Create contact string with name and email
                contact = f"{user_dict['name']} ({user_dict['email']})" if user_dict['name'].strip() else user_dict['email']
                
                # Create a list for the current row
                row = [
                    user_dict['organization'],
                    contact,
                    f"{user_dict['city']}, {user_dict['state']}",
                    "",
                    "",
                    ""
                ]
                
                # Add the row to the data list
                sheet.append(row)

            for row in sheet['A1':'F'+ str(len(users_list)+1)]:
                for cell in row:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            adjust_width(sheet)


            program_labels_workbook.create_sheet(title="Sign In Day 2")
            program_labels_workbook.active = program_labels_workbook["Sign In Day 2"]
            sheet = program_labels_workbook.active

            # list of column headers
            headers = ["Program/School", "Contact", "Location", "Registration Packet", "T-shirts/Bags", "Posters"]

            # set the background color of the first row to blue, for the number of columnes equal to the length of the headers list
            for cell in sheet['A1':'F1']:
                for c in cell:
                    c.font = Font(bold=True)

            # set the values of the second row to the column headers
            for i, header in enumerate(headers):
                sheet.cell(row=1, column=i+1, value=header)

            # filter the users_list to only include users that are in submissions that are on day 2
            users_list_day2 = [user_dict for user_dict in users_list if user_dict['day2'] == "Day 2"]
            for user_dict in users_list_day2:
                # Create contact string with name and email
                contact = f"{user_dict['name']} ({user_dict['email']})" if user_dict['name'].strip() else user_dict['email']
                
                # Create a list for the current row
                row = [
                    user_dict['organization'],
                    contact,
                    f"{user_dict['city']}, {user_dict['state']}",
                    "",
                    "",
                    ""
                ]
                
                # Add the row to the data list
                sheet.append(row)

            for row in sheet['A1':'F'+ str(len(users_list)+1)]:
                for cell in row:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            adjust_width(sheet)


            # remove the default sheet
            program_labels_workbook.remove(program_labels_workbook['Sheet'])

            # Create a BytesIO object and save the workbook to it
            xlsx_file_io = BytesIO()
            program_labels_workbook.save(xlsx_file_io)

            # Go back to the start of the BytesIO object
            xlsx_file_io.seek(0)

            # Save the BytesIO object to a file in default storage
            program_labels_xlsx_file_name = f'Fair{fair.name}-Program labels.xlsx'
            program_labels_xlsx_file = default_storage.save(program_labels_xlsx_file_name, ContentFile(xlsx_file_io.read()))


            # Get the full path of the files
            json_file_path = default_storage.path(json_file)
            submission_xlsx_file_path = default_storage.path(submission_xlsx_file)
            student_xlsx_file_path = default_storage.path(student_xlsx_file)
            group_contact_xlsx_path = default_storage.path(group_contact_xlsx_file)
            accessory_xlsx_file_path = default_storage.path(accessory_xlsx_file)
            program_labels_file_path = default_storage.path(program_labels_xlsx_file)

            zip_folder_name = f'fair_{fair.name}_data/'

            # Create a new zip file
            zip_file_name = f'fair_{fair.name}_data.zip'
            with zipfile.ZipFile(zip_file_name, 'w') as zip_file:
                # Add the json file to the zip file
                zip_file.write(json_file_path, arcname=zip_folder_name+json_file_name)

                # Add the submission xlsx file to the zip file
                zip_file.write(submission_xlsx_file_path, arcname=zip_folder_name+submission_xlsx_file_name)

                # Add the student xlsx file to the zip file
                zip_file.write(student_xlsx_file_path, arcname=zip_folder_name+student_xlsx_file_name)

                # Add the group contact xlsx file to the zip file
                zip_file.write(group_contact_xlsx_path, arcname=zip_folder_name+group_contact_xlsx_file_name)

                # Add the accessory xlsx file to the zip file
                zip_file.write(accessory_xlsx_file_path, arcname=zip_folder_name+accessory_xlsx_file_name)

                # Add the program labels xlsx file to the zip file
                zip_file.write(program_labels_file_path, arcname=zip_folder_name+program_labels_xlsx_file_name)

            # Create a generator that reads the file and yields the data
            def file_iterator():
                try:
                    with open(zip_file_name, 'rb') as f:
                        for chunk in iter(lambda: f.read(4096), b''):
                            yield chunk
                finally:
                    os.remove(zip_file_name)

            # Create a StreamingHttpResponse that uses the file_iterator
            response = StreamingHttpResponse(file_iterator(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{zip_file_name}"'

            # Delete the files from the default storage
            default_storage.delete(json_file)
            default_storage.delete(submission_xlsx_file)
            default_storage.delete(student_xlsx_file)
            default_storage.delete(group_contact_xlsx_file)
            default_storage.delete(accessory_xlsx_file)
            default_storage.delete(program_labels_xlsx_file)

            return response

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# and API view that returns JSON for all the submissions for the fair given by the fair_pk, with all the metadata for each submission. This is sent to the browser as a download when the user clicks the "Download All Submission data" button on the fair detail page.
class JudgeSheetsDownloadView(APIView):
    def get(self, request, fair_pk):
        fair = Fair.objects.get(pk=fair_pk)
        submissions = Submission.objects.filter(fair=fair)

        ## Make xlsx for submissions (non material)

        # filter submissions to only include those that are approved
        submissions = submissions.filter(status__in=["approved"])

        # filter submissions by category, exluding the categories "Poster", "Comics and Cartoons", "Mobile Video"
        submissions = submissions.exclude(category__name__in=["Poster", "Comics and Cartoons", "Mobile Video"])

        # sort submissions by category, then by organization, then by grade range, then by group, then by title
        submissions = submissions.order_by('category__name', 'organization', 'grade_range', 'group', 'title')

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="fair{fair.name}-Judging sheets.pdf"'
        
        p = reportlab_canvas.Canvas(response, pagesize=reportlab_letter)
        p.setTitle(f'Fair {fair.name} - Judging Sheets')
        width, height = reportlab_letter

        # Register the AboriginalSansREGULAR font
        font_path = os.path.join(settings.STATIC_ROOT, 'AboriginalSansREGULAR.ttf')
        # font_bold_path = os.path.join(settings.STATIC_ROOT, 'AboriginalSansBOLD.ttf')
        reportlab_pdfmetrics.registerFont(reportlab_TTFont('AboriginalSansREGULAR', font_path))
        # reportlab_pdfmetrics.registerFont(reportlab_TTFont('AboriginalSansBOLD', font_bold_path))

        def draw_static_elements():
            # Title banner image
            # Get image
            image_path = os.path.join(settings.STATIC_ROOT, 'onaylf.png')
            # Create an Image object
            image = reportlab_Image(image_path)
            # Set the width to match the page width (with margins)
            image_width = width - 60  # 30px margin on each side
            # Calculate height to maintain aspect ratio
            aspect = image.imageHeight / float(image.imageWidth)
            image_height = image_width * aspect
            # Resize the image
            image.drawWidth = image_width
            image.drawHeight = image_height
            # Draw the image
            image.drawOn(p, 30, height-100)
            # p.drawString(30, height-30, "SAM NOBLE MUSEUM DEPARTMENT OF NATIVE AMERICAN LANGUAGES")
            # Subtitle with Blue Background
            sky_blue = reportlab_Color(0.429, 0.708, 0.982)  # RGB values for light blue
            p.setFillColor(sky_blue)
            rectangle_width = width * 0.8  # 80% of the page width
            rectangle_x = (width - rectangle_width) / 2  # Calculate the x-coordinate to center the rectangle

            p.rect(rectangle_x, height-140, rectangle_width, 30, fill=True, stroke=False)
            p.setFillColor(reportlab_white)
            p.setFont("Helvetica", 14)
            p.drawString(130, height-130, f"Oklahoma Native American Youth Language Fair {fair.name}")

            p.setFillColor(reportlab_black)
            # Horizontal Rule after Dynamic Text
            p.line(30, height-300, width-30, height-300)

            p.setFont("Helvetica", 9)
            p.drawString(35, height-315, f"Please rate the {submission.category.name}")
            p.drawString(35, height-325, "using a scale of 1 to 5.")
            p.drawString(435, height-315, "Circle score for each")
            p.drawString(435, height-325, "critera:")
            p.setFont("Helvetica-Bold", 9)
            p.drawString(35, height-345, "1 = Poor")
            p.drawString(35, height-355, "2 = Below Average")
            p.drawString(35, height-365, "3 = Average")
            p.drawString(35, height-375, "4 = Above Average")
            p.drawString(35, height-385, "5 = Excellent")

            # Likert-style Ratings
            ratings = ["                                                   Use of Language:",
                       "Originality & Creativity in Illustration & Content:",
                       "                                                                       Effort:"]
            y_position = height - 345  # Subtract 100 from the y-coordinate
            for rating in ratings:
                p.drawString(220, y_position, rating)
                for i in range(1, 6):
                    p.drawString(420 + 15*i, y_position, str(i))
                y_position -= 20

            # Horizontal Rule after Likerts
            p.line(30, y_position-10, width-30, y_position-10)

            # Total Score Box, Comments Box, Judge Name and Signature
            p.drawString(350, y_position-20, "Total Score:")
            p.rect(420, y_position-40, 100, 30)
            p.drawString(50, y_position-60, "Comments:")
            p.rect(48, y_position-225, width-100, 150)
            p.drawString(130, y_position-280, "Judge Name:")
            p.line(200, y_position-280, width-130, y_position-280)
            p.drawString(130, y_position-320, "Signature:")
            p.line(200, y_position-320, width-130, y_position-320)
            

            # Horizontal Rule before footer
            p.line(30, 40, width-30, 40)
            # Footer
            p.setFont("Helvetica", 8)
            p.drawString(30, 30, f"Oklahoma Native American Youth Language Fair {fair.name} Judging Sheet - {submission.category.name}")

        for submission in submissions:
            draw_static_elements()

            # Dynamic text drawing goes here
            p.setFont("AboriginalSansREGULAR", 10)
            p.drawString(30, height-155, "Program/School: ")
            p.drawString(150, height-155, f"{submission.organization}")
            p.drawString(30, height-175, "Grade: ")
            p.drawString(150, height-175, f"{submission.get_grade_range_display()}")
            p.drawString(30, height-195, "Presenting Group: ")
            p.drawString(150, height-195, f"{submission.group}")
            p.drawString(30, height-215, "Title: ")
            p.drawString(150, height-215, f"{submission.title}")
            p.drawString(30, height-235, "Language: ")
            languoid_text = []
            for languoid in submission.languoids.all():
                if languoid.name == 'Other':
                    other_text = f"Other: {submission.other_languoid or 'Blank'}"
                    languoid_text.append(other_text)
                else:
                    languoid_text.append(languoid.name)
            p.drawString(150, height-235, f"{', '.join(languoid_text)}")
            p.drawString(30, height-255, "Category: ")
            p.drawString(150, height-255, f"{submission.category.name}")
            p.drawString(30, height-275, "Type: ")
            p.drawString(150, height-275, f"{submission.get_submission_type_display()}")

            p.showPage()

        p.save()
        return response

# and API view that returns JSON for all the submissions for the fair given by the fair_pk, with all the metadata for each submission. This is sent to the browser as a download when the user clicks the "Download All Submission data" button on the fair detail page.
class SubmissionSheetsDownloadView(APIView):
    def get(self, request, fair_pk):
        fair = Fair.objects.get(pk=fair_pk)
        submissions = Submission.objects.filter(fair=fair)

        # filter submissions to only include those that are approved
        submissions = submissions.filter(status__in=["approved"])

        # filter submissions to include only those that are not material
        submissions = submissions.filter(category__material_submission=False)

        # # filter submissions by category, exluding the categories "Poster", "Comics and Cartoons", "Mobile Video"
        # submissions = submissions.exclude(category__name__in=["Poster", "Comics and Cartoons", "Mobile Video"])

        # sort submissions by category, then by organization, then by grade range, then by group, then by title
        submissions = submissions.order_by('category__name', 'organization', 'grade_range', 'group', 'title')
        
        # Define an onPage function that draws the footer
        def footer(canvas, doc):
            canvas.saveState()
            pageNumber = canvas.getPageNumber()
            total_pages = doc.page
            canvas.setFont('AboriginalSansREGULAR', 10)
            canvas.drawCentredString(page_width / 2, 20, f"Page {pageNumber}")
            canvas.restoreState()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="fair{fair.name}-Submission sheets.pdf"'
        doc = reportlab_SimpleDocTemplate(response, pagesize=reportlab_letter, onPage=footer, title=f'Fair {fair.name} - Submission Sheets')

        styles = reportlab_getSampleStyleSheet()
        styles['Normal'].fontName = 'AboriginalSansREGULAR'
        styles['Heading2'].fontName = 'AboriginalSansREGULAR'

        # Create a new style based on 'Normal'
        font_bigger_bold = reportlab_ParagraphStyle('font_bigger_bold', parent=styles['Normal'], fontSize=13, bold=True, italic=False)
        # Add the new style to the stylesheet
        styles.add(font_bigger_bold)

        # # Define a Frame for the main content
        # frame_main = reportlab_Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='main')

        # # Define a PageTemplate that uses this Frame
        # template = reportlab_PageTemplate(id='test', frames=[frame_main])

        # # Apply this PageTemplate to the document
        # doc.addPageTemplates([template])

        elements = []
        page_width, page_height = reportlab_letter

        # Register the AboriginalSansREGULAR font
        font_path = os.path.join(settings.STATIC_ROOT, 'AboriginalSansREGULAR.ttf')
        font_bold_path = os.path.join(settings.STATIC_ROOT, 'AboriginalSansBOLD.ttf')
        reportlab_pdfmetrics.registerFont(reportlab_TTFont('AboriginalSansREGULAR', font_path))
        reportlab_pdfmetrics.registerFont(reportlab_TTFont('AboriginalSansBOLD', font_bold_path))

        for submission in submissions:
            # add the header image onaylf.png
            image_path = os.path.join(settings.STATIC_ROOT, 'onaylf.png')
            image_width = page_width - 200
            image = reportlab_Image(image_path, width=image_width, height=60)
            elements.append(image)

            # Add submission details
            clean_title = escape(submission.title)
            elements.append(reportlab_Paragraph(f"<b>Title of Presentation:</b> {clean_title}", styles['Heading2']))
            elements.append(reportlab_Paragraph(f"<b>Presenting Group Name:</b> {submission.group}", styles['Normal']))
            # add some horizontal space
            elements.append(reportlab_Spacer(1, 2))
            elements.append(reportlab_Paragraph(f"<b>School/Program:</b> {submission.organization}", styles['Normal']))
            elements.append(reportlab_Spacer(1, 1))
            data = [[
                reportlab_Paragraph(f"<b>Submission Type:</b> {submission.submission_type}", styles['Normal']),
                reportlab_Paragraph(f"<b>Category:</b> {submission.category}", styles['Normal'])
            ]]
            table = reportlab_Table(data)
            table.setStyle(reportlab_TableStyle([
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                ('GRID', (0,0), (-1,-1), 1, (0,0,0,0)),  # Set table border color to transparent
                ('LEFTPADDING', (0,0), (-1,-1), 0),  # Remove left padding
            ]))
            elements.append(table)
            # add a text line that is the count of the unique students in the submission
            elements.append(reportlab_Paragraph(f"<b>Student count:</b> {submission.students.count()}", styles['Normal']))

            # add some horizontal space
            elements.append(reportlab_Spacer(1, 12))
            # add a horizontal line break
            elements.append(reportlab_HRFlowable(width="100%", thickness=1, lineCap='round', color=reportlab_Color(0, 0, 0)))
            # add some horizontal space
            elements.append(reportlab_Spacer(1, 10))

            total_width = page_width - 150  # Total width of the table
            # Calculate the widths of the columns as percentages of the total width
            year_width = total_width * 0.2
            grade_width = total_width * 0.3 
            language_width = total_width * 0.5

            languoid_text = []
            for languoid in submission.languoids.all():
                if languoid.name == 'Other':
                    other_text = f"Other: {submission.other_languoid or 'Blank'}"  # Add 'Blank' as default
                    languoid_text.append(other_text)
                else:
                    languoid_text.append(languoid.name)

            data = [[
                reportlab_Paragraph(f"<b>Year:</b> {fair.name}", styles['font_bigger_bold']),
                reportlab_Paragraph(f"<b>Grade:</b> {submission.get_grade_range_display()}", styles['font_bigger_bold']),
                reportlab_Paragraph(f"<b>Language(s):</b> {', '.join(languoid_text)}", styles['font_bigger_bold'])  # New element
            ]]
            table = reportlab_Table(data, colWidths=[year_width, grade_width, language_width])
            table.setStyle(reportlab_TableStyle([
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                ('GRID', (0,0), (-1,-1), 1, (0,0,0,0)),  # Set table border color to transparent
                ('LEFTPADDING', (0,0), (-1,-1), 0),  # Remove left padding
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),  # Add bottom padding
                ('TOPPADDING', (0,0), (-1,-1), 6),  # Add top padding
            ]))
            elements.append(table)

            # add some horizontal space
            elements.append(reportlab_Spacer(1, 10))
            # add a horizontal line break
            elements.append(reportlab_HRFlowable(width="100%", thickness=1, lineCap='round', color=reportlab_Color(0, 0, 0)))
            # add some horizontal space
            elements.append(reportlab_Spacer(1, 12))

            # Dynamic table for instructors
            instructor_data = [[
                reportlab_Paragraph("<b>First Name</b>", styles['Normal']), 
                reportlab_Paragraph("<b>Last Name</b>", styles['Normal'])
            ]]  # Table header
            instructors = Instructor.objects.filter(submission_instructor=submission)
            for instructor in instructors:
                instructor_data.append([
                    reportlab_Paragraph(instructor.firstname, styles['Normal']),
                    reportlab_Paragraph(instructor.lastname, styles['Normal'])
                ])
            instructor_table = reportlab_Table(instructor_data, colWidths=[150, 150])
            instructor_table.setStyle(reportlab_TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), reportlab_colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, reportlab_colors.black),
            ]))

            # Add title
            elements.append(reportlab_Paragraph("<b>Instructors</b>", styles['Heading2']))
            elements.append(instructor_table)

            # Dynamic table for students
            student_data = [[
                reportlab_Paragraph("<b>First Name</b>", styles['Normal']), 
                reportlab_Paragraph("<b>Last Name</b>", styles['Normal'])
            ]]  # Table header
            students = Student.objects.filter(submission_student=submission)
            for student in students:
                student_data.append([
                    reportlab_Paragraph(student.firstname, styles['Normal']),
                    reportlab_Paragraph(student.lastname, styles['Normal'])
                ])
            student_table = reportlab_Table(student_data, colWidths=[150, 150])
            student_table.setStyle(reportlab_TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), reportlab_colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, reportlab_colors.black),
            ]))
            elements.append(reportlab_Spacer(1, 12))  # Add some space between tables
            elements.append(reportlab_Paragraph("<b>Students</b>", styles['Heading2']))
            elements.append(student_table)
            elements.append(reportlab_PageBreak())  # Ensure each submission starts on a new page

        doc.build(elements, onFirstPage=footer, onLaterPages=footer)

        return response

# and API view that returns pdf for organizations with summary info
class RegistrationCoverSheetsDownloadView(APIView):
    def get(self, request, fair_pk):
        fair = Fair.objects.get(pk=fair_pk)

        # get a list of cateogries that are not material
        non_material_submission_categories = list(Category.objects.filter(fair=fair, material_submission=False).values_list('name', flat=True))

        # get a list of all the approved submissions for the fair
        submissions = Submission.objects.filter(fair=fair, status__in=["approved"])

        # get a list of all the users that have students that are in the submissions of the fair
        users = User.objects.filter(student_user__submission_student__in=submissions).distinct()

        # Convert GRADES to a dictionary
        grades_dict = dict(Student.GRADES)

        # create a mapping from display values to sorted non-display values (this should have been done in models)
        tshirt_size_mapping = {
            'Youth Small (YS)': '1_ys',
            'Youth Medium (YM)': '2_ym',
            'Youth Large (YL)': '3_yl',
            'Adult Small (S)': '4_s',
            'Adult Medium (M)': '5_m',
            'Adult Large (L)': '6_l',
            'Adult Extra Large (XL)': '7_xl',
            'Adult Extra Extra Large (XXL)': '8_xxl',
            'Adult Extra Extra Extra Large (XXXL)': '9_xxxl'
        }

        # make a list of dictionaries for each user, with the organization of the user, address, city and state of the user, phone, fax and email of the user
        users_list = []
        for user in users:

            # make a clean organization name, that strips the characters that are escaped from the beginning of the organization name
            if user.organization is not None:
                clean_organization = re.sub(r'^\W+', '', user.organization)
            else:
                clean_organization = ""

            # get the submissions that the user is in
            submissions = Submission.objects.filter(user=user, fair=fair)

            # get the students that are in the submissions, sorted by last name and then first name
            students = Student.objects.filter(submission_student__in=submissions).order_by('lastname', 'firstname').distinct()

            # get a list of cateogries that are not material
            non_material_categories = Category.objects.filter(fair=fair, material_submission=False)

            # make a list of dictionaries for each student, with the first name, last name, grade, and tshirt size of the student
            students_list = []
            for student in students:
                # get the submissions that the student is in
                student_submissions = student.submission_student.filter(fair=fair)

                # check if the student is in any non-material submissions
                has_non_material = student_submissions.filter(category__in=non_material_categories).exists()

                # Only include tshirt size if student has non-material submissions
                tshirt_size = student.get_tshirt_size_display() if has_non_material else ""

                # create a dictionary for the student
                student_dict = {
                    'firstname': student.firstname,
                    'lastname': student.lastname,
                    'grade': student.get_grade_display(),
                    'tshirt_size': tshirt_size
                }
                # add a copy of the dictionary to the list of students
                students_list.append(copy.deepcopy(student_dict))

            # get the count of the unique students in the submissions
            student_count = students.count()

            # using students_list, get a count of the tshirts of each size as a dictionary
            tshirt_sizes = Counter([student['tshirt_size'] for student in students_list])

            # remove the count for blank t-shirt sizes and set it to bag_count
            bag_count = tshirt_sizes.pop('', 0)


            # sort tshirt_sizes based on the non-display value of the vocabulary
            tshirt_sizes = {k: v for k, v in sorted(tshirt_sizes.items(), key=lambda item: tshirt_size_mapping[item[0]])}

            # get the instructors that are in the submissions
            instructors = Instructor.objects.filter(submission_instructor__in=submissions).distinct()

            # get the count of the unique instructors in the submissions
            instructor_count = instructors.count()

            # get the grades of the students
            grades = sorted(list(set(students.values_list('grade', flat=True))))

            # for each value in grades, replace with the display value from grades_dict
            grades = [grades_dict[grade] for grade in grades]

            # create a dictionary for the user
            user_dict = {
                'first_name': user.first_name,
                'last_name': user.last_name,
                'organization': clean_organization,
                'address': user.address,
                'city_state_zip': ", ".join(filter(None, [user.city, " ".join(filter(None, [user.state, user.zip]))])),
                'phone': ", ".join(filter(None, [user.phone, user.alt_phone])),
                'fax': user.fax or '',
                'email': ', '.join(filter(None, [user.email, user.alt_email])) or '',
                'grades': ", ".join([str(grade) for grade in grades]),
                'students': students_list,
                'student_count': student_count,
                'tshirt_sizes': tshirt_sizes,
                'bag_count': bag_count,
                'instructors': instructors,
                'instructor_count': instructor_count
            }
            # add a copy of the dictionary to the list of users
            users_list.append(copy.deepcopy(user_dict))

        # sort the list of users by the organization name, ignoring None values
        users_list = sorted(users_list, key=lambda x: (x['organization'] is None, x['organization'].lower() if x['organization'] else ''))

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="fair{fair.name}-Registration cover sheets.pdf"'
        doc = reportlab_SimpleDocTemplate(response, pagesize=reportlab_letter, title=f'Fair {fair.name} - Registration Cover Sheets')

        # Register the AboriginalSansREGULAR font
        font_path = os.path.join(settings.STATIC_ROOT, 'AboriginalSansREGULAR.ttf')
        font_bold_path = os.path.join(settings.STATIC_ROOT, 'AboriginalSansBOLD.ttf')
        reportlab_pdfmetrics.registerFont(reportlab_TTFont('AboriginalSansREGULAR', font_path))
        reportlab_pdfmetrics.registerFont(reportlab_TTFont('AboriginalSansBOLD', font_bold_path))

        styles = reportlab_getSampleStyleSheet()
        styles['Normal'].fontName = 'AboriginalSansREGULAR'
        styles['Heading2'].fontName = 'AboriginalSansREGULAR'

        # Create a new style based on 'Normal'
        font_bold = reportlab_ParagraphStyle('font_bold', parent=styles['Normal'], bold=True, italic=False)
        # Add the new style to the stylesheet
        styles.add(font_bold)

        # Create a new style based on 'Normal'
        font_bigger_bold = reportlab_ParagraphStyle('font_bigger_bold', parent=styles['Normal'], fontSize=13, bold=True, italic=False)
        # Add the new style to the stylesheet
        styles.add(font_bigger_bold)

        elements = []
        page_width, page_height = reportlab_letter

        # add the header image onaylf.png
        image_path = os.path.join(settings.STATIC_ROOT, 'onaylf.png')
        image_width = page_width - 200
        image = reportlab_Image(image_path, width=image_width, height=60)

        for user in users_list:
            # Add the banner image at the start of each user's section
            elements.append(image)

            # Add organization details
            
            total_width = page_width - 155  # Total width of the table
            # Calculate the widths of the columns as percentages of the total width
            name_width = total_width * 0.9
            letter_width = total_width * 0.1

            data = [[
                reportlab_Paragraph(f"{user['organization']}", styles['Heading1']),
                reportlab_Paragraph(f"{user['organization'][0] if user['organization'] else ''}", styles['Heading1']),
            ]]
            table = reportlab_Table(data)
            table = reportlab_Table(data, colWidths=[name_width, letter_width])
            table.setStyle(reportlab_TableStyle([
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                ('GRID', (0,0), (-1,-1), 1, (0,0,0,0)),  # Set table border color to transparent
                ('LEFTPADDING', (0,0), (-1,-1), 0),  # Remove left padding
            ]))
            elements.append(table)

            # add a horizontal line break
            elements.append(reportlab_HRFlowable(width="100%", thickness=1, lineCap='round', color=reportlab_Color(0, 0, 0)))
            # add some horizontal space
            elements.append(reportlab_Spacer(1, 4))

            elements.append(reportlab_Paragraph(f"Registration Cover Sheet - {fair.name}", styles['Normal']))
            # add some horizontal space
            elements.append(reportlab_Spacer(1, 2))

            total_width = page_width - 155  # Total width of the table
            # Calculate the widths of the columns as percentages of the total width
            label_width = total_width * 0.2
            value_width = total_width * 0.8

            # For the instructors table
            data = [[
                reportlab_Paragraph("Instructor(s)", styles['Normal']),
                reportlab_Paragraph(", ".join([f'{instructor.firstname} {instructor.lastname}' for instructor in user['instructors']]), styles['Normal'])
            ]]
            table = reportlab_Table(data, colWidths=[label_width, value_width])
            table.setStyle(reportlab_TableStyle([
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                ('GRID', (0,0), (-1,-1), 1, (0,0,0,0)),  # Set table border color to transparent
                ('LEFTPADDING', (0,0), (-1,-1), 0),  # Remove left padding
                ('VALIGN', (0, 0), (-1, -1), 'TOP')  # Add this line for top alignment
            ]))
            elements.append(table)

            # For the contact table
            contact_lines = []
            
            print(user)

            # Name and email line (required)
            name = " ".join(filter(None, [user.get('first_name', ''), user.get('last_name', '')]))
            if name and user.get('email'):
                contact_lines.append(f"{name} ({user.get('email')})")
            elif name:
                contact_lines.append(name)
            else:
                contact_lines.append(user.get('email', ''))  # At minimum, show email

            # Address line (optional)
            address_parts = filter(None, [user.get('address', ''), user.get('city_state_zip', '')])
            address = "<br />".join(address_parts)
            if address:
                contact_lines.append(address)

            # Phone/Fax line (optional)
            phone_fax_parts = []
            if user.get('phone'):
                phone_fax_parts.append(user.get('phone'))
            if user.get('fax'):
                phone_fax_parts.append(f"{user.get('fax')} (fax)")
            if phone_fax_parts:
                contact_lines.append(" , ".join(phone_fax_parts))

            # Create the contact table (always)
            data = [[
                reportlab_Paragraph("Contact", styles['Normal']),
                reportlab_Paragraph("<br />".join(contact_lines), styles['Normal'])
            ]]
            table = reportlab_Table(data, colWidths=[label_width, value_width])
            table.setStyle(reportlab_TableStyle([
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                ('GRID', (0,0), (-1,-1), 1, (0,0,0,0)),  # Set table border color to transparent
                ('LEFTPADDING', (0,0), (-1,-1), 0),  # Remove left padding
                ('VALIGN', (0, 0), (-1, -1), 'TOP')  # Add this line for top alignment
            ]))
            elements.append(table)

            # add some horizontal space
            elements.append(reportlab_Spacer(1, 4))
            # add a horizontal line break
            elements.append(reportlab_HRFlowable(width="100%", thickness=1, lineCap='round', color=reportlab_Color(0, 0, 0)))
            # add some horizontal space
            elements.append(reportlab_Spacer(1, 4))

            elements.append(reportlab_Paragraph("Total Attendees", styles['font_bigger_bold']))
            # add some horizontal space
            elements.append(reportlab_Spacer(1, 8))
            elements.append(reportlab_Paragraph(f"Students:   {user['student_count']}", styles['Normal']))
            elements.append(reportlab_Paragraph(f"Instructors: {user['instructor_count']}", styles['Normal']))

            # add some horizontal space
            elements.append(reportlab_Spacer(1, 12))

            elements.append(reportlab_Paragraph("Grades in Program/School", styles['font_bigger_bold']))
            # add some horizontal space
            elements.append(reportlab_Spacer(1, 8))
            elements.append(reportlab_Paragraph(f"{user['grades']}", styles['Normal']))

            # add some horizontal space
            elements.append(reportlab_Spacer(1, 12))

            elements.append(reportlab_Paragraph("Number of Bags", styles['font_bigger_bold']))
            # add some horizontal space
            elements.append(reportlab_Spacer(1, 8))
            elements.append(reportlab_Paragraph(f"{user['bag_count']}", styles['Normal']))

            # make a table of tshirt sizes, showing the size and the count of each size
            tshirt_data = [["T-Shirt Size", "Count"]]
            for size, count in user['tshirt_sizes'].items():
                tshirt_data.append([size, count])
            tshirt_table = reportlab_Table(tshirt_data, colWidths=[175, 100])
            tshirt_table.setStyle(reportlab_TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), reportlab_colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, reportlab_colors.black),
            ]))
            elements.append(reportlab_Spacer(1, 12))  # Add some space between tables
            elements.append(reportlab_Paragraph("T-Shirt Sizes", styles['font_bigger_bold']))
            elements.append(tshirt_table)

            # add some horizontal space
            elements.append(reportlab_Spacer(1, 12))

            # Dynamic table for students
            student_data = [["First Name", "Last Name", "Grade", "T-shirt Size", "Flag"]]  # Table header
            for student in user['students']:
                student_data.append([student['firstname'], student['lastname'], student['grade'], student['tshirt_size'], ""])
            student_table = reportlab_Table(student_data, colWidths=[110, 110, 80, 150, 50])
            student_table.setStyle(reportlab_TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), reportlab_colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, reportlab_colors.black),
            ]))
            elements.append(reportlab_Spacer(1, 12))  # Add some space between tables
            elements.append(reportlab_Paragraph("<b>Students</b>", styles['Heading2']))
            elements.append(student_table)
            elements.append(reportlab_PageBreak())  # Ensure each submission starts on a new page

        doc.build(elements)

        return response

# and API view that returns JSON for all the submissions for the fair given by the fair_pk, with all the metadata for each submission. This is sent to the browser as a download when the user clicks the "Download All Submission data" button on the fair detail page.
class SubmissionCardsDownloadView(APIView):
    def get(self, request, fair_pk):
        fair = Fair.objects.get(pk=fair_pk)
        submissions = Submission.objects.filter(fair=fair)

        # filter submissions to only include those that are approved
        submissions = submissions.filter(status__in=["approved"])

        # filter submissions to only include those that are not material
        submissions = submissions.filter(category__material_submission=False)

        # # filter submissions by category, exluding the categories "Poster", "Comics and Cartoons", "Mobile Video"
        # submissions = submissions.exclude(category__name__in=["Poster", "Comics and Cartoons", "Mobile Video"])

        # sort submissions by category, then by organization, then by grade range, then by group, then by title
        submissions = submissions.order_by('category__name', 'organization', 'grade_range', 'group', 'title')

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="fair{fair.name}-Submission cards.pdf"'
        
        p = reportlab_canvas.Canvas(response, pagesize=reportlab_letter)
        p.setTitle(f'Fair {fair.name} - Submission Cards')
        width, height = reportlab_letter

        # Register the AboriginalSansREGULAR font
        font_path = os.path.join(settings.STATIC_ROOT, 'AboriginalSansREGULAR.ttf')
        font_bold_path = os.path.join(settings.STATIC_ROOT, 'AboriginalSansBOLD.ttf')
        reportlab_pdfmetrics.registerFont(reportlab_TTFont('AboriginalSansREGULAR', font_path))
        reportlab_pdfmetrics.registerFont(reportlab_TTFont('AboriginalSansBOLD', font_bold_path))

        # Title banner image
        # Get image
        image_path = os.path.join(settings.STATIC_ROOT, 'onaylf.png')
        # Create an Image object
        image = reportlab_Image(image_path)
        # Set the size of the image
        image.drawWidth = 330
        image.drawHeight = 50

        def draw_static_elements():
            p.line(30, height-398, width-30, height-398)

        def draw_dynamic_elements(y_position, submission):
            image.drawOn(p, 30, y_position+30)
            p.setFont("AboriginalSansBOLD", 14)
            p.drawString(380, y_position+60, "Submission card")

            # Dynamic text drawing goes here
            p.setFont("AboriginalSansREGULAR", 10)
            p.drawString(50, y_position, "Presenting Group: ")
            p.drawString(80, y_position-20, f"{submission.group}")
            p.drawString(350, y_position, "Category: ")
            p.drawString(380, y_position-20, f"{submission.category.name}")
            p.drawString(50, y_position-60, "Program/School: ")
            p.drawString(80, y_position-80, f"{(submission.organization or '')[:50]}")
            p.drawString(350, y_position-60, "Grade: ")
            p.drawString(380, y_position-80, f"{submission.get_grade_range_display()}")
            p.drawString(50, y_position-120, "Title of Presentation: ")
            p.drawString(80, y_position-140, f"{(submission.title or '')[:50]}")
            p.drawString(350, y_position-120, "Type: ")
            p.drawString(380, y_position-140, f"{submission.get_submission_type_display()}")
            p.drawString(50, y_position-180, "Language(s): ")
            languoid_text = []
            for languoid in submission.languoids.all():
                if languoid.name == 'Other':
                    other_text = f"Other: {submission.other_languoid or 'Blank'}"
                    languoid_text.append(other_text)
                else:
                    languoid_text.append(languoid.name)
            p.drawString(80, y_position-200, f"{', '.join(languoid_text)}")
            p.drawString(350, y_position-180, "Instructor(s): ")
            p.drawString(380, y_position-200, f"{', '.join([instructor.lastname for instructor in submission.instructors.all()])}")


        for i, submission in enumerate(submissions):
            if i % 2:
                pass
            else:
                try:
                    second_submission = submissions[i+1]
                except:
                    second_submission = None

                draw_static_elements()
                y_position = height-130
                draw_dynamic_elements(y_position, submission)
                y_position = height-530
                if second_submission:
                    draw_dynamic_elements(y_position, second_submission)

                p.showPage()

        p.save()
        return response


# def query_inveniordm(request):
#     # The URL to the InvenioRDM records API endpoint
#     url = "https://143.244.215.98/api/records"

#     # Parameters for the search query
#     params = {
#         "q": "metadata.languages.props.family:(\"coch1271\")"
#     } 

# # https://143.244.215.98/api/records?q=%22rood%22%20AND%20metadata.languages.id%3A%28%22wich1260%22%29


#     # Make a GET request to the API
#     try:
#         response = requests.get(url, params=params, verify=False)  # Use verify=False only if necessary for self-signed certificates
#     except requests.exceptions.RequestException as e:
#         # Handle request exceptions (e.g., network errors)
#         return JsonResponse({"error": "Network error or invalid URL", "details": str(e)}, status=500)

#     # Check if the request was successful
#     if response.status_code == 200:
#         # Parse the response JSON and return it
#         data = response.json()
#         return JsonResponse(data, safe=False)
#     else:
#         # Handle errors or unsuccessful responses
#         error_message = response.text
#         try:
#             error_data = response.json()
#             error_message = error_data.get('message', response.text)
#         except ValueError:
#             pass  # JSON parsing failed, use the raw text
#         return JsonResponse({"error": "Failed to fetch data", "details": error_message}, status=response.status_code)






# submissions_all = Submission.objects.filter(status__in=["approved", "submitted"])

# submissions_non = submissions_all.filter(Q(category__name="Master Performer") | Q(category__name="Modern Song") | Q(category__name="Skit/Short Play") | Q(category__name="Spoken Language") | Q(category__name="Spoken Poetry") | Q(category__name="Spoken Prayer") | Q(category__name="Traditional Song"))

# students_non = Student.objects.filter(submission_student__in=submissions_non).distinct()


# submissions_material = submissions_all.filter(Q(category__name="Books") | Q(category__name="Comics and Cartoons") | Q(category__name="Film and Video") | Q(category__name="Mobile Video") | Q(category__name="Poster") | Q(category__name="Puppet Show"))

# students_material = Student.objects.filter(submission_student__in=submissions_material).distinct()

# # find the students that are in both submissions and submissions_material
# students_both = students_non.filter(id__in=students_material)

@user_passes_test(is_admin)
def migrate_data(request):
    # Get all available fairs
    fairs = Fair.objects.all()
    
    if request.method == 'POST':
        if 'preview' in request.POST:
            # Handle file upload and preview
            if 'json_file' not in request.FILES:
                messages.error(request, 'Please select a file to upload')
                return redirect('migrate_data')
            
            file = request.FILES['json_file']
            try:
                data = json.load(file)
            except json.JSONDecodeError:
                messages.error(request, 'Invalid JSON file')
                return redirect('migrate_data')

            # Process and organize data for preview
            preview_data = {
                'users': {},
                'submissions': [],
                'students': {},
                'instructors': {}
            }

            existing_users = set(User.objects.values_list('email', flat=True))

            # Process users with default values and missing field tracking
            for item in data:
                user_data = item['user']
                email = user_data['email']
                
                if email not in preview_data['users']:
                    # Track which fields came from JSON
                    from_json = ['id', 'email', 'first_name', 'last_name']
                    missing = []
                    
                    # Check optional fields
                    optional_fields = [
                        'organization', 'phone', 'alt_phone', 'fax', 
                        'alt_email', 'address', 'city', 'state', 'zip'
                    ]
                    
                    for field in optional_fields:
                        if field in user_data:
                            from_json.append(field)
                        else:
                            missing.append(field)
                            user_data[field] = None
                    
                    # Add default values for required fields
                    user_data.update({
                        'exists': email in existing_users,
                        'date_joined': timezone.now().isoformat(),
                        'is_staff': False,
                        'is_active': True,
                        # Track field categories
                        'from_json': from_json,
                        'defaulted': ['date_joined', 'is_staff', 'is_active'],
                        'missing': missing
                    })
                    preview_data['users'][email] = user_data

                # Track submission (rename performance_type to submission_type)
                submission = item.copy()
                submission['submission_type'] = submission.pop('performance_type', '')
                
                # Add organization from user data
                submission['organization'] = item['user'].get('organization', '')
                
                # Keep references to associated students and instructors
                submission['student_ids'] = [s['id'] for s in submission['students']]
                submission['instructor_ids'] = [i['id'] for i in submission['instructors']]
                # Store accessories with their counts - with more defensive checks
                submission['accessory_data'] = []
                for acc in submission.get('accessories', []):
                    try:
                        acc_data = {
                            'id': acc['accessory']['id'],
                            'name': acc['accessory']['name'],
                            'count': acc.get('count', 0)  # Default to 0 if count not present
                        }
                        submission['accessory_data'].append(acc_data)
                    except (KeyError, TypeError) as e:
                        print(f"Warning: Malformed accessory data: {acc}")
                        continue
    
                
                preview_data['submissions'].append(submission)
                
                # Track students and instructors with their IDs
                for student in item['students']:
                    if student['id'] not in preview_data['students']:
                        preview_data['students'][student['id']] = student

                for instructor in item['instructors']:
                    if instructor['id'] not in preview_data['instructors']:
                        preview_data['instructors'][instructor['id']] = instructor

                # Track accessories with their IDs
                for accessory_data in item.get('accessories', []):
                    acc = accessory_data['accessory']  # Get the nested accessory object
                    if acc['id'] not in preview_data.get('accessories', {}):
                        preview_data.setdefault('accessories', {})[acc['id']] = acc

            # Convert dictionaries to lists for template rendering
            preview_data['users'] = list(preview_data['users'].values())
            preview_data['students'] = list(preview_data['students'].values())
            preview_data['instructors'] = list(preview_data['instructors'].values())

            # Store preview data in session for later use
            request.session['migration_preview'] = preview_data
            return render(request, 'migrate.html', {'preview': preview_data, 'fairs': fairs})

        elif 'confirm' in request.POST:
            preview_data = request.session.get('migration_preview')
            if not preview_data:
                messages.error(request, 'No preview data found')
                return redirect('migrate_data')

            # Validate fair selection
            fair_id = request.POST.get('fair')
            if not fair_id:
                messages.error(request, 'Please select a fair')
                return render(request, 'migrate.html', {
                    'preview': preview_data,
                    'fairs': fairs
                })

            try:
                fair = Fair.objects.get(id=fair_id)
            except Fair.DoesNotExist:
                messages.error(request, 'Invalid fair selected')
                return render(request, 'migrate.html', {
                    'preview': preview_data,
                    'fairs': fairs
                })

            try:
                with transaction.atomic():  # Use transaction to ensure data consistency
                    logger.info("Starting data migration...")
                    
                    # Create users
                    logger.info("Creating users...")
                    existing_users = set(User.objects.values_list('email', flat=True))
                    for user_data in preview_data['users']:
                        logger.info(f"Processing user: {user_data['email']}")
                        if user_data['email'] not in existing_users:
                            user = User.objects.create(
                                email=user_data['email'],
                                first_name=user_data['first_name'],
                                last_name=user_data['last_name'],
                                organization=user_data.get('organization'),  # Add organization
                                phone=user_data.get('phone'),
                                alt_phone=user_data.get('alt_phone'),
                                fax=user_data.get('fax'),
                                alt_email=user_data.get('alt_email'),
                                address=user_data.get('address'),
                                city=user_data.get('city'),
                                state=user_data.get('state'),
                                zip=user_data.get('zip'),
                                date_joined=timezone.now(),
                                is_staff=False,
                                is_active=True
                            )

                    # Create students
                    logger.info("Creating students...")
                    student_id_map = {}  # Old ID -> New Student object
                    for student_data in preview_data['students']:
                        logger.info(f"Processing student: {student_data['firstname']} {student_data['lastname']}")
                        creator = User.objects.get(email=student_data['user']['email'])
                        student = Student.objects.create(
                            user=creator,
                            fair=fair,
                            firstname=student_data['firstname'],
                            lastname=student_data['lastname'],
                            grade=student_data['grade'],
                            hometown=student_data['hometown'],
                            state=student_data['state'],
                            tshirt_size=student_data['tshirt_size']
                        )
                        student_id_map[student_data['id']] = student

                    # Create instructors
                    logger.info("Creating instructors...")
                    instructor_id_map = {}  # Old ID -> New Instructor object
                    for instructor_data in preview_data['instructors']:
                        logger.info(f"Processing instructor: {instructor_data['firstname']} {instructor_data['lastname']}")
                        creator = User.objects.get(email=instructor_data['user']['email'])
                        instructor = Instructor.objects.create(
                            user=creator,
                            fair=fair,
                            firstname=instructor_data['firstname'],
                            lastname=instructor_data['lastname']
                        )
                        instructor_id_map[instructor_data['id']] = instructor

                    # Create submissions and link relationships
                    logger.info("Creating submissions...")
                    for submission_data in preview_data['submissions']:
                        logger.info(f"Processing submission: {submission_data['title']}")
                        creator = User.objects.get(email=submission_data['user']['email'])
                        submission = Submission.objects.create(
                            user=creator,
                            fair=fair,
                            organization=creator.organization or '',
                            title=submission_data['title'],
                            group=submission_data['group'],
                            category=Category.objects.get(name=submission_data['category']['name']),
                            grade_range=submission_data['grade_range'],
                            submission_type=submission_data['submission_type'],
                            status=submission_data['status']
                        )

                        # Handle languoids
                        for lang_name in submission_data.get('languoid', []):
                            try:
                                languoid = Languoid.objects.get(name=lang_name)
                                submission.languoids.add(languoid)
                            except Languoid.DoesNotExist:
                                logger.warning(f"Languoid '{lang_name}' not found for submission '{submission.title}'")

                        # Add associated students using the ID map
                        for student_id in submission_data['student_ids']:
                            if student_id in student_id_map:
                                submission.students.add(student_id_map[student_id])

                        # Add associated instructors using the ID map
                        for instructor_id in submission_data['instructor_ids']:
                            if instructor_id in instructor_id_map:
                                submission.instructors.add(instructor_id_map[instructor_id])

                        # Add associated accessories with counts
                        for acc_data in submission_data['accessory_data']:
                            accessory = Accessory.objects.get(id=acc_data['id'])
                            SubmissionAccessory.objects.create(
                                submission=submission,
                                accessory=accessory,
                                count=acc_data['count']
                            )

                    logger.info("Data migration completed successfully")
                    messages.success(request, 'Data migration completed successfully')
                    return redirect('home')

            except Exception as e:
                logger.error(f"Error during migration: {str(e)}", exc_info=True)
                messages.error(request, f'Error during migration: {str(e)}')
                return redirect('migrate_data')

    return render(request, 'migrate.html', {'fairs': fairs})

@require_http_methods(["DELETE"])
@login_required
@user_passes_test(is_moderator)
def submission_delete(request, submission_id):
    if request.method != 'DELETE':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
        
    try:
        submission = get_object_or_404(Submission, id=submission_id)
        submission.delete()
        return JsonResponse({'status': 'success'})
    except Exception as e:
        logger.error(f"Error deleting submission {submission_id}: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@require_http_methods(["GET"])
def check_student_delete(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    
    # Check for submissions associated with this student
    submissions = Submission.objects.filter(students=student)
    
    if submissions.exists():
        # Student has associated submissions - cannot delete
        return JsonResponse({
            'can_delete': False,
            'submissions': [{
                'id': sub.id,
                'title': sub.title,
                'category': sub.category.name
            } for sub in submissions]
        })
    else:
        # No associations - safe to delete
        return JsonResponse({
            'can_delete': True,
            'submissions': []
        })

@login_required
@require_http_methods(["DELETE"])
def delete_student(request, student_id):
    try:
        student = get_object_or_404(Student, id=student_id)
        
        # Check for submissions before deleting
        if student.submission_student.exists():
            return JsonResponse({
                'error': 'Cannot delete student with associated submissions'
            }, status=400)
        
        student.delete()
        return JsonResponse({'success': True})
        
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=400)


import json, os, re
from io import BytesIO
import zipfile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.colors import blue, white, black, Color
from reportlab.platypus import Image
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from django.db.models import Min, Max, Sum
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.generic.edit import FormView, DeleteView
from django.core.mail import send_mail
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.conf import settings
from django.http import Http404, HttpResponse, StreamingHttpResponse
from rest_framework import generics, viewsets, status
from rest_framework.views import APIView
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from users.models import User
from .models import STATE_CHOICES, Fair, CurrentFair, Languoid, Tribe, Performance, Category, Instructor, Student, Accessory, PerformanceAccessory
from .serializers import CategorySerializer, PerformanceSerializer, PosterSerializer, InstructorSerializer, StudentSerializer, PerformanceAccessorySerializer, PerformanceJsonSerializer
from .forms import PerformanceForm, PerformanceCommentsForm, InstructorForm, StudentForm, PosterForm

import requests
from django.http import JsonResponse
import logging

logger = logging.getLogger(__name__)

def is_moderator(user):
    return user.groups.filter(name='moderator').exists()

def custom_500_view(request):
    return render(request, "500.html", status=500)

def contact_info(request):
    currentFair = CurrentFair.objects.first()

    template = 'contact_info.html'
    context = {
        'currentFair': currentFair.name,
    }
    return render(request, template, context)

@api_view(['GET'])
def performance_list(request):
    if not request.user.is_authenticated:
        return Response({'detail': 'Authentication credentials were not provided.'}, status=status.HTTP_401_UNAUTHORIZED)
    performances = Performance.objects.filter(poster=False)
    user_id = request.GET.get('user_id')
    if user_id is not None:
        user = get_object_or_404(User, pk=user_id)
        performances = performances.filter(user=user)
    serializer = PerformanceSerializer(performances, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def poster_list(request):
    if not request.user.is_authenticated:
        return Response({'detail': 'Authentication credentials were not provided.'}, status=status.HTTP_401_UNAUTHORIZED)
    performances = Performance.objects.filter(poster=True)
    serializer = PosterSerializer(performances, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def performance_poster_list(request):
    if not request.user.is_authenticated:
        return Response({'detail': 'Authentication credentials were not provided.'}, status=status.HTTP_401_UNAUTHORIZED)
    performances = Performance.objects.all()
    user_id = request.GET.get('user_id')
    if user_id is not None:
        user = User.objects.get(pk=user_id)
        performances = performances.filter(user=user)
    serializer = PerformanceSerializer(performances, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def performance_get(request, perf_pk):
    if not request.user.is_authenticated:
        return Response({'detail': 'Authentication credentials were not provided.'}, status=status.HTTP_401_UNAUTHORIZED)
        # Check if the user is in the "moderator" group
    performance = get_object_or_404(Performance, pk=perf_pk)
    if not request.user.groups.filter(name='moderator').exists():
        if not request.user == performance.user:
            return Response({'detail': 'You do not have permission to perform this action.'}, status=status.HTTP_403_FORBIDDEN)
    serializer = PerformanceSerializer(performance)
    return Response(serializer.data)

class CategoryUpdateView(LoginRequiredMixin, generics.UpdateAPIView):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    # permission_classes = [IsAuthenticated]  # Example: Allow any user to update tasks !!!!!!!!!!!!!!!

class InstructorAddView(APIView):
    def post(self, request):
        # Get the current fair
        currentFair = CurrentFair.objects.first()
        # Get the user id from the request data
        user_id = request.data.get('user_id')

        # Get the user from the request
        user = User.objects.get(id=user_id)

        # Create a new instructor with the data from the request
        instructor = Instructor(
            firstname=request.data.get('firstname'),
            lastname=request.data.get('lastname'),
            user=user,
            fair=currentFair.fair,
            modified_by=request.user.get_username(),
        )

        # Save the instructor
        instructor.save()

        # Serialize the instructor
        serializer = InstructorSerializer(instructor)

        # Return the serialized instructor
        return Response(serializer.data, status=status.HTTP_201_CREATED)

class InstructorUpdateView(APIView):
    def put(self, request, instr_pk):
        instructor = self.get_object(instr_pk)
        serializer = InstructorSerializer(instructor, data=request.data)
        if serializer.is_valid():
            serializer.validated_data['modified_by'] = request.user.get_username()
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get_object(self, instr_pk):
        try:
            return Instructor.objects.get(pk=instr_pk)
        except Instructor.DoesNotExist:
            raise Http404

class InstructorViewSet(LoginRequiredMixin, viewsets.ModelViewSet):
    serializer_class = InstructorSerializer

    def get_queryset(self):
        queryset = Instructor.objects.all()
        user_id = self.request.query_params.get('user_id', None)
        performance_id = self.request.query_params.get('performance_id', None)
        if user_id is not None:
            queryset = queryset.filter(user__id=user_id)
        if performance_id is not None:
            queryset = queryset.filter(performance_instructor__id=performance_id)
        return queryset

class StudentAddView(APIView):
    def post(self, request):
        # Get the current fair
        currentFair = CurrentFair.objects.first()
        # Get the user id from the request data
        user_id = request.data.get('user_id')

        # Get the user from the request
        user = User.objects.get(id=user_id)

        # Create a new student with the data from the request
        student = Student(
            firstname=request.data.get('firstname'),
            lastname=request.data.get('lastname'),
            grade=request.data.get('grade'),
            hometown=request.data.get('hometown'),
            state=request.data.get('state'),
            tshirt_size=request.data.get('tshirt_size'),
            user=user,
            fair=currentFair.fair,
            modified_by=request.user.get_username(),
        )

        # Save the student
        student.save()

        student.tribe.set(request.data.get('tribes'))

        # Serialize the student
        serializer = StudentSerializer(student)

        # Return the serialized student
        return Response(serializer.data, status=status.HTTP_201_CREATED)

class StudentUpdateView(APIView):
    def put(self, request, stud_pk):
        # Get the current fair
        currentFair = CurrentFair.objects.first()
        # Get the user id from the request data
        user_id = request.data.get('user_id')

        # Get the user from the request
        user = User.objects.get(id=user_id)

        # Get the student to update
        student = get_object_or_404(Student, pk=stud_pk)

        # Update the student with the data from the request
        student.firstname = request.data.get('firstname')
        student.lastname = request.data.get('lastname')
        student.grade = request.data.get('grade')
        student.hometown = request.data.get('hometown')
        student.state = request.data.get('state')
        student.tshirt_size = request.data.get('tshirt_size')
        # student.user = user # This line is commented out because the user should not be updated
        # student.fair = currentFair.fair # This line is commented out because the fair should not be updated
        student.modified_by = request.user.get_username()

        # Save the student
        student.save()

        student.tribe.set(request.data.get('tribes'))

        # Serialize the student
        serializer = StudentSerializer(student)

        # Return the serialized student
        return Response(serializer.data, status=status.HTTP_200_OK)


class StudentViewSet(LoginRequiredMixin, viewsets.ModelViewSet):
    serializer_class = StudentSerializer

    def get_queryset(self):
        queryset = Student.objects.all()
        user_id = self.request.query_params.get('user_id', None)
        performance_id = self.request.query_params.get('performance_id', None)
        if user_id is not None:
            queryset = queryset.filter(user__id=user_id)
        if performance_id is not None:
            queryset = queryset.filter(performance_student__id=performance_id)
        return queryset

class PerformanceUpdateView(LoginRequiredMixin, generics.UpdateAPIView):
    queryset = Performance.objects.all()
    serializer_class = PerformanceSerializer

class PerformanceAccessoryCreateView(LoginRequiredMixin, generics.CreateAPIView):
    queryset = PerformanceAccessory.objects.all()
    serializer_class = PerformanceAccessorySerializer

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def perform_create(self, serializer):
        serializer.save()

class PerformanceAccessoryUpdateView(LoginRequiredMixin, generics.UpdateAPIView):
    queryset = PerformanceAccessory.objects.all()
    serializer_class = PerformanceAccessorySerializer

    def get_object(self):
        performance_id = self.kwargs.get('perf_pk')
        accessory_id = self.kwargs.get('acc_pk')
        return get_object_or_404(PerformanceAccessory, performance=performance_id, accessory=accessory_id)

@login_required
def select_fair(request, pk=None):

    currentFair = CurrentFair.objects.first()

    if pk:
        get_fair_to_update = Fair.objects.get(id=pk)
        if get_fair_to_update:
            currentFair.fair = get_fair_to_update
            currentFair.name = get_fair_to_update.name
            currentFair.save()

    fairs = Fair.objects.values('id', 'name')
    fairs_ordered = fairs.order_by('-updated')

    recent_fairs = fairs.order_by('-updated') # this modification of fairs is currently redundant, but it would be needed if the ordering of fairs above needed to change
    if len(recent_fairs) > 3:
        recent_fairs = recent_fairs[:3]

    template = 'select_fair.html'
    context = {
        'currentFair': currentFair.name,
        'recent_fairs': recent_fairs,
        'fairs': fairs_ordered
    }
    return render(request, template, context)


class PerformanceAccessoryViewSet(LoginRequiredMixin, viewsets.ModelViewSet):
    queryset = PerformanceAccessory.objects.all()
    serializer_class = PerformanceAccessorySerializer

    def get_queryset(self):
        queryset = PerformanceAccessory.objects.all()
        performance_id = self.request.query_params.get('performance_id', None)
        accessory_id = self.request.query_params.get('accessory_id', None)

        if performance_id is not None:
            queryset = queryset.filter(performance__id=performance_id)
        if accessory_id is not None:
            queryset = queryset.filter(accessory__id=accessory_id)
        return queryset

@login_required
def edit_fair(request, pk):

    currentUserEmail = request.user.get_username()

    currentUser = User.objects.get(email=currentUserEmail)

    is_moderator = currentUser.groups.filter(name='moderator').exists()

    currentFair = CurrentFair.objects.first()

    fair = Fair.objects.prefetch_related("fair_categories", "fair_accessories").get(id=pk)

    print(fair.fair_categories)
    template = 'edit_fair.html'
    context = {
        'moderator': is_moderator,
        'currentFair': currentFair.name,
        'fair': fair,
    }
    return render(request, template, context)

@login_required
def home(request):

    currentUserEmail = request.user.get_username()

    currentUser = User.objects.get(email=currentUserEmail)

    # Check if the user is a moderator
    is_moderator = currentUser.groups.filter(name='moderator').exists()
    print(is_moderator)

    currentFair = CurrentFair.objects.first()

    performances = Performance.objects.prefetch_related("user", "category").filter(fair=currentFair.fair)

    if not is_moderator:
        print("this is working")
        print(performances)
        performances = performances.filter(user=currentUser)
        print(performances)

    template = 'home.html'
    context = {
        'currentUser': currentUser,
        'moderator': is_moderator,
        'registrationOpen': currentFair.fair.registration_open,
        'currentFair': currentFair.name,
        'performances': performances
    }
    return render(request, template, context)


@login_required
@user_passes_test(is_moderator)
def user_list(request):

    currentFair = CurrentFair.objects.first()

    # Check if the user is a moderator
    is_moderator = request.user.groups.filter(name='moderator').exists()

    # get all users sorted by first name
    all_users = User.objects.order_by('first_name')

    template = 'user_list.html'
    context = {
        'moderator': is_moderator,
        'currentFair': currentFair.name,
        'users': all_users
    }
    return render(request, template, context)


@login_required
@user_passes_test(is_moderator)
def user_detail(request, user_pk):

    currentUser = User.objects.get(pk=user_pk)

    currentFair = CurrentFair.objects.first()

    # Check if the user is a moderator
    is_moderator = request.user.groups.filter(name='moderator').exists()

    allPerformances = Performance.objects.prefetch_related("user", "students").filter(fair=currentFair.fair).filter(user=currentUser.pk)

    performances = allPerformances.filter(poster=False)

    posters = allPerformances.filter(poster=True)

    template = 'user_detail.html'
    context = {
        'currentUser': currentUser,
        'moderator': is_moderator,
        'currentFair': currentFair.name,
        'performances': performances,
        'posters': posters
    }
    return render(request, template, context)

@login_required
def performance_detail(request, perf_pk):

    currentFair = CurrentFair.objects.first()

    # Check if the user is a moderator
    is_moderator = request.user.groups.filter(name='moderator').exists()

    performance = Performance.objects.prefetch_related("instructors", "students", "accessories").get(pk=perf_pk)

    performance_user_organization = performance.user.organization

    # get the non material submission categories for the current fair
    non_material_submission_categories = list(Category.objects.filter(fair=currentFair.fair, material_submission=False).values_list('name', flat=True))
    # check if this performance is a non-material submission
    non_material_submission = performance.category.name in non_material_submission_categories

    # Fetch the PerformanceAccessory instances related to the current performance
    performance_accessories = PerformanceAccessory.objects.filter(performance=performance)

    # Create a dictionary mapping accessory IDs to counts
    accessory_counts = {pa.accessory.id: pa.count for pa in performance_accessories}

    # Fetch the accessories
    accessories = Accessory.objects.filter(fair=currentFair.fair)

    # Update the count attribute of each accessory with the count from the performance_accessories
    for accessory in accessories:
        if accessory.id in accessory_counts:
            accessory.count = accessory_counts[accessory.id]

    other_languoid = Languoid.objects.get(name='Other')

    performance_includes_other_languoid = performance.languoids.filter(pk=other_languoid.pk).exists()

    template = 'performance_detail.html'
    context = {
        'currentFair': currentFair.name,
        'moderator': is_moderator,
        'performance': performance,
        'organization': performance_user_organization,
        'non_material_submission': non_material_submission,
        'includes_other_languoid': performance_includes_other_languoid,
        'accessories': accessories

    }
    return render(request, template, context)

class performance_add(LoginRequiredMixin, FormView):
    def handle_no_permission(self):
        return redirect('/accounts/login/')
    form_class = PerformanceForm
    template_name = "performance_add.html"
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['add_type'] = 'Performance'

        if 'user_pk' in self.kwargs:
            user_pk = self.kwargs['user_pk']  # Access the pk value
            user = User.objects.get(pk=user_pk)  # Get the user object
        else:
            user = self.request.user
        context['owning_user'] = user

        currentFair = CurrentFair.objects.first()
        context['currentFair'] = currentFair.name

        context['selected_category'] = self.kwargs.get('category', None)


        material_submission_categories = list(Category.objects.filter(fair=currentFair.fair, material_submission=True).values_list('name', flat=True))
        context['material_submission_categories'] = json.dumps(material_submission_categories)

        non_material_submission_categories = list(Category.objects.filter(fair=currentFair.fair, material_submission=False).values_list('name', flat=True))
        context['non_material_submission_categories'] = json.dumps(non_material_submission_categories)

        categories = Category.objects.filter(fair=currentFair.fair)
        categories_list = list(categories.values('name', 'max_students'))
        context['categories'] = json.dumps(categories_list)

        categories = Category.objects.filter(fair=currentFair.fair)
        categories_dict = {category.name: category.max_students for category in categories}
        context['category_student_max'] = json.dumps(categories_dict)

        accessories = Accessory.objects.filter(fair=currentFair.fair)
        context['accessories'] = accessories

        grades = Student.GRADES
        context['grades'] = grades

        tribes = Tribe.objects.filter(fair=currentFair.fair)
        context['tribes'] = tribes

        states = STATE_CHOICES
        context['states'] = states

        tshirt_sizes = Student.TSHIRT_SIZES
        context['tshirt_sizes'] = tshirt_sizes

        return context

    def get(self, request, *args, **kwargs):
        form = self.form_class(selected_category=self.get_context_data()['selected_category'])
        return self.render_to_response(self.get_context_data(form=form))
    
    def form_valid(self, form):
        if form.is_valid():
            currentFair = CurrentFair.objects.first()
            self.object = form.save(commit=False)
            self.object.fair = currentFair.fair
            if 'user_pk' in self.kwargs:
                user_pk = self.kwargs['user_pk']  # Access the pk value
                user = User.objects.get(pk=user_pk)  # Get the user object
                self.object.user = user
            else:
                self.object.user = self.request.user
            self.object.modified_by = self.request.user.get_username()
            self.object.save()
            form.save_m2m()

            # Get the instructors from the form data
            instructors_json = self.request.POST.get('instructors')

            # Check if instructors_json exists
            if instructors_json:
                # Parse the instructors_json as a JSON array
                instructor_ids = json.loads(instructors_json)

                # For each instructor_id in the instructor_ids array...
                for instructor_id in instructor_ids:
                    # Get the Instructor object with the given ID
                    instructor = Instructor.objects.get(id=instructor_id)

                    # Add the instructor to the performance's instructors
                    self.object.instructors.add(instructor)

            # Get the students from the form data
            students_json = self.request.POST.get('students')

            # Check if students_json exists
            if students_json:
                # Parse the students_json as a JSON array
                student_ids = json.loads(students_json)

                # For each student_id in the student_ids array...
                for student_id in student_ids:
                    # Get the Student object with the given ID
                    student = Student.objects.get(id=student_id)

                    # Add the student to the performance's students
                    self.object.students.add(student)

            # Set performance_type based on the numner of students
            if self.object.students.count() > 1:
                self.object.performance_type = "group"
            else:
                self.object.performance_type = "individual"

            # get the student in the highest grade
            max_grade = self.object.students.aggregate(Max('grade'))
            max_grade_value = max_grade['grade__max']
            # look up the grade range based on the highest grade
            grade_range = Performance.GRADE_RANGES_DICT.get(max_grade_value)
            # set the grade range (based on the student in the highest grade)
            self.object.grade_range = grade_range

            # Save the performance object to update the instructors and students
            self.object.save()

            # Get the performance_accessory_counts from the form data
            performance_accessory_counts = self.request.POST.get('performance_accessory_counts')

            # Check if performance_accessory_counts exists
            if performance_accessory_counts:
                # Parse the performance_accessory_counts as a JSON object
                try:
                    accessory_counts = json.loads(performance_accessory_counts)
                except json.JSONDecodeError:
                    print(f"Invalid JSON: {performance_accessory_counts}")
                    accessory_counts = {}
                # For each key-value pair in the accessory_counts object...
                for accessory_id, count in accessory_counts.items():
                    # Create a PerformanceAccessory object
                    PerformanceAccessory.objects.create(
                        performance=self.object,
                        accessory_id=accessory_id,
                        count=count
                    )

            self.object.instructors_status = "completed"
            self.object.students_status = "completed"
            self.object.accessories_status = "completed"
            self.object.review_status = "completed"
            self.object.status = "submitted"
            self.object.save()

            # Determine the redirect URL based on the request path
            print("submit-and-add" in self.request.POST)
            if 'submit-and-add' in self.request.POST:
                return redirect("/performance/add/")
            else:
                return redirect("/performance/%s/" % self.object.pk)
    def form_invalid(self, form):
        print(form.errors)
        return super().form_invalid(form)

class performance_add_admin(UserPassesTestMixin, performance_add):
    def test_func(self):
        return self.request.user.groups.filter(name='moderator').exists()
    
@login_required
def performance_instructors(request, perf_pk):

    currentFair = CurrentFair.objects.first()

    performance = Performance.objects.get(pk=perf_pk)

    if performance.instructors_status == "pending":
        performance.instructors_status = "in_progress"
        performance.save()

    template = 'performance_instructors.html'
    context = {
        'currentFair': currentFair.name,
        'performance': performance
    }
    return render(request, template, context)

class instructor_add(LoginRequiredMixin, FormView):
    def handle_no_permission(self):
        return redirect('/no-permission')
    form_class = InstructorForm
    template_name = "instructor_add.html"
    def form_valid(self, form):
        if form.is_valid():
            currentFair = CurrentFair.objects.first()
            self.object = form.save(commit=False)
            self.object.fair = currentFair.fair
            if 'perf_pk' in self.kwargs:
                performance_id = self.kwargs['perf_pk']  # Access the pk value
                performance = Performance.objects.get(id=performance_id)
                self.object.user = performance.user
            else:
                self.object.user = self.request.user
            self.object.modified_by = self.request.user.get_username()
            self.object.save()
            # Determine the redirect URL based on the request path
            if 'performance/' in self.request.path:
                # return redirect(reverse('performance:performance_instructors_add', kwargs={'perf_pk': self.kwargs['perf_pk']}))
                return redirect("../")
            else:
                return redirect("../../")
    def form_invalid(self, form):
        print(form.errors)
        return super().form_invalid(form)

@login_required    
def instructor_edit(request, instr_pk, perf_pk=None):

    currentFair = CurrentFair.objects.first()

    instructor = get_object_or_404(Instructor, id=instr_pk)

    if request.method == "POST":
        form = InstructorForm(request.POST, instance=instructor)
        if form.is_valid():
            form.save(commit=False)
            form.modified_by = request.user.get_username()
            form.save()
            if perf_pk:
                return redirect("../../")
            else:
                return redirect("../../../")
    else:
        form = InstructorForm(instance=instructor)

    template = 'instructor_edit.html'
    context = {
        'currentFair': currentFair.name,
        'form': form
    }
    return render(request, template, context)

@login_required
def performance_students(request, perf_pk):

    currentFair = CurrentFair.objects.first()

    performance = Performance.objects.get(pk=perf_pk)

    if performance.students_status == "pending":
        performance.students_status = "in_progress"
        performance.save()

    template = 'performance_students.html'
    context = {
        'currentFair': currentFair.name,
        'performance': performance
    }
    return render(request, template, context)

class student_add(LoginRequiredMixin, FormView):
    def handle_no_permission(self):
        return redirect('/no-permission')
    form_class = StudentForm
    template_name = "student_add.html"
    def form_valid(self, form):
        if form.is_valid():
            currentFair = CurrentFair.objects.first()
            self.object = form.save(commit=False)
            self.object.fair = currentFair.fair
            if 'perf_pk' in self.kwargs:
                performance_id = self.kwargs['perf_pk']  # Access the pk value
                performance = Performance.objects.get(id=performance_id)
                self.object.user = performance.user
            else:
                self.object.user = self.request.user
            self.object.modified_by = self.request.user.get_username()
            self.object.save()
            # Determine the redirect URL based on the request path
            if 'performance/' in self.request.path:
                # return redirect(reverse('performance:performance_students_add', kwargs={'perf_pk': self.kwargs['perf_pk']}))
                return redirect("../")
            else:
                return redirect("../../")

            return redirect("../")
    def form_invalid(self, form):
        print(form.errors)
        return super().form_invalid(form)

@login_required
def student_edit(request, stud_pk, perf_pk=None):

    currentFair = CurrentFair.objects.first()

    student = get_object_or_404(Student, id=stud_pk)

    if request.method == "POST":
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save(commit=False)
            form.modified_by = request.user.get_username()
            form.save()
            return redirect("../../")
    else:
        form = StudentForm(instance=student)

    template = 'student_edit.html'
    context = {
        'currentFair': currentFair.name,
        'form': form
    }
    return render(request, template, context)

@login_required
def performance_accessories(request, perf_pk):

    currentFair = CurrentFair.objects.first()

    performance = Performance.objects.get(pk=perf_pk)

    if performance.accessories_status == "pending":
        performance.accessories_status = "in_progress"
        performance.save()

    # Fetch the PerformanceAccessory instances related to the current performance
    performance_accessories = PerformanceAccessory.objects.filter(performance=performance)

    # Create a dictionary mapping accessory IDs to counts
    accessory_counts = {pa.accessory.id: pa.count for pa in performance_accessories}

    # Fetch the accessories
    accessories = Accessory.objects.filter(fair=currentFair.fair)

    # Update the count attribute of each accessory with the count from the performance_accessories
    for accessory in accessories:
        if accessory.id in accessory_counts:
            accessory.count = accessory_counts[accessory.id]

    template = 'performance_accessories.html'
    context = {
        'currentFair': currentFair.name,
        'performance': performance,
        'accessories': accessories
    }
    return render(request, template, context)

@login_required
def performance_edit(request, perf_pk):

    # Check if the user is a moderator
    is_moderator = request.user.groups.filter(name='moderator').exists()

    currentFair = CurrentFair.objects.first()

    performance = get_object_or_404(Performance, id=perf_pk)

    # Redirect if performance status is not 'in_progress'
    if performance.status != 'in_progress':
        return redirect('performance_detail', perf_pk=perf_pk)

    material_submission_categories = list(Category.objects.filter(fair=currentFair.fair, material_submission=True).values_list('name', flat=True))
    non_material_submission_categories = list(Category.objects.filter(fair=currentFair.fair, material_submission=False).values_list('name', flat=True))

    categories = Category.objects.filter(fair=currentFair.fair)
    categories_list = list(categories.values('name', 'max_students'))
    categories_dict = {category.name: category.max_students for category in categories}

    other_languoid = Languoid.objects.get(name='Other')

    performance_includes_other_languoid = performance.languoids.filter(pk=other_languoid.pk).exists()

    accessories = Accessory.objects.filter(fair=currentFair.fair)

    grades = Student.GRADES

    tribes = Tribe.objects.filter(fair=currentFair.fair)

    states = STATE_CHOICES

    tshirt_sizes = Student.TSHIRT_SIZES

    if request.method == "POST":
        form = PerformanceForm(request.POST, instance=performance)
        if form.is_valid():
            edited_performance = form.save(commit=False)
            edited_performance.modified_by = request.user.get_username()
            edited_performance.save()
            form.save_m2m()

            # Get the instructors from the form data
            instructors_json = request.POST.get('instructors')

            # Check if instructors_json exists
            if instructors_json:
                # Parse the instructors_json as a JSON array
                instructor_ids = json.loads(instructors_json)

                # Get the Instructor objects with the given IDs
                instructors = Instructor.objects.filter(id__in=instructor_ids)

                # Set the performance's instructors to match the provided list
                edited_performance.instructors.set(instructors)

            # Get the students from the form data
            students_json = request.POST.get('students')

            # Check if students_json exists
            if students_json:
                # Parse the students_json as a JSON array
                student_ids = json.loads(students_json)

                # Get the Student objects with the given IDs
                students = Student.objects.filter(id__in=student_ids)

                # Set the performance's students to match the provided list
                edited_performance.students.set(students)

            # if override_performance_type is checked, do nothing special with performance type
            if not edited_performance.override_performance_type:
            # otherwise: set performance_type based on the numner of students
                if edited_performance.students.count() > 1:
                    edited_performance.performance_type = "group"
                else:
                    edited_performance.performance_type = "individual"


            # get the student in the highest grade
            max_grade = edited_performance.students.aggregate(Max('grade'))
            max_grade_value = max_grade['grade__max']
            # look up the grade range based on the highest grade
            grade_range = Performance.GRADE_RANGES_DICT.get(max_grade_value)
            # set the grade range (based on the student in the highest grade)
            edited_performance.grade_range = grade_range

            # Save the performance object to update the instructors and students
            edited_performance.save()

            # Get the performance_accessory_counts from the form data
            performance_accessory_counts = request.POST.get('performance_accessory_counts')

            # Check if performance_accessory_counts exists
            if performance_accessory_counts:
                # Parse the performance_accessory_counts as a JSON object
                try:
                    accessory_counts = json.loads(performance_accessory_counts)
                except json.JSONDecodeError:
                    print(f"Invalid JSON: {performance_accessory_counts}")
                    accessory_counts = {}
                # For each key-value pair in the accessory_counts object...
                for accessory_id, count in accessory_counts.items():
                    # Create a PerformanceAccessory object
                    PerformanceAccessory.objects.update_or_create(
                        performance=edited_performance,
                        accessory_id=accessory_id,
                        defaults={'count': count}
                    )
            
            edited_performance.instructors_status = "completed"
            edited_performance.students_status = "completed"
            edited_performance.accessories_status = "completed"
            edited_performance.review_status = "completed"
            edited_performance.status = "submitted"
            edited_performance.save()

            return redirect(".")
    else:
        form = PerformanceForm(instance=performance)
    template = 'performance_add.html'
    context = {
        'moderator': is_moderator,
        'currentFair': currentFair.name,
        'performance': performance,
        'owning_user': performance.user,
        'material_submission_categories': json.dumps(material_submission_categories),
        'non_material_submission_categories': json.dumps(non_material_submission_categories),
        'categories': json.dumps(categories_list),
        'category_student_max': json.dumps(categories_dict),
        'includes_other_languoid': performance_includes_other_languoid,
        'accessories': accessories,
        'grades': grades,
        'tribes': tribes,
        'states': states,
        'tshirt_sizes': tshirt_sizes,
        'form': form
    }
    return render(request, template, context)

@login_required
def performance_review(request, perf_pk):

    currentFair = CurrentFair.objects.first()

    performance = Performance.objects.prefetch_related("instructors", "students", "accessories").get(pk=perf_pk)

    if performance.review_status == "pending" :
        performance.review_status = "in_progress"
        performance.save()

    performance_user_organization = performance.user.organization

    # Fetch the PerformanceAccessory instances related to the current performance
    performance_accessories = PerformanceAccessory.objects.filter(performance=performance)

    # Create a dictionary mapping accessory IDs to counts
    accessory_counts = {pa.accessory.id: pa.count for pa in performance_accessories}

    # Fetch the accessories
    accessories = Accessory.objects.filter(fair=currentFair.fair)

    # Update the count attribute of each accessory with the count from the performance_accessories
    for accessory in accessories:
        if accessory.id in accessory_counts:
            accessory.count = accessory_counts[accessory.id]
    
    other_languoid = Languoid.objects.get(name='Other')

    performance_includes_other_languoid = performance.languoids.filter(pk=other_languoid.pk).exists()

    if request.method == "POST":
        form = PerformanceCommentsForm(request.POST, instance=performance)
        if form.is_valid():
            print("valid penis")
            performance_form = form.save(commit=False)
            performance_form.review_status = "completed"
            performance_form.modified_by = request.user.get_username()
            performance_form.save()
            return redirect("/")
    else:
        form = PerformanceCommentsForm(instance=performance)

    template = 'performance_review.html'
    context = {
        'currentFair': currentFair.name,
        'performance': performance,
        'includes_other_languoid': performance_includes_other_languoid,
        'form': form
    }
    return render(request, template, context)

class poster_add(LoginRequiredMixin, FormView):
    def handle_no_permission(self):
        return redirect('/no-permission')
    form_class = PosterForm
    template_name = "poster_add.html"
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if 'user_pk' in self.kwargs:
            user_pk = self.kwargs['user_pk']  # Access the pk value
            user = User.objects.get(pk=user_pk)  # Get the user object
        else:
            user = self.request.user
        context['owning_user'] = user
        return context
    def form_valid(self, form):
        if form.is_valid():
            if 'user_pk' in self.kwargs:
                user_pk = self.kwargs['user_pk']  # Access the pk value
                user = User.objects.get(pk=user_pk)  # Get the user object
            else:
                user = self.request.user
            currentFair = CurrentFair.objects.first()
            self.object = form.save(commit=False)
            self.object.fair = currentFair.fair
            self.object.poster = True
            self.object.user = user
            self.object.instructors_status = "completed"
            self.object.students_status = "completed"
            self.object.accessories_status = "completed"
            self.object.review_status = "completed"
            self.object.modified_by = self.request.user.get_username()
            self.object.save()
            # self.object.title = "Poster " + str(self.object.pk)  # Set the name here
            self.object.save()  # Save the object again to store the new name
            form.save_m2m()
            for performance_accessory_count in form.performance_accessory_counts:
                if performance_accessory_count['count'] > 0:
                    print(performance_accessory_count['accessory'])
                    performance_accessory = PerformanceAccessory(performance=self.object, accessory=performance_accessory_count['accessory'], count=performance_accessory_count['count'])
                    performance_accessory.save()
            return redirect("../../")
    def form_invalid(self, form):
        print(form.errors)
        return super().form_invalid(form)


class poster_add_admin(UserPassesTestMixin, poster_add):
    def test_func(self):
        return self.request.user.groups.filter(name='moderator').exists()

@login_required
def poster_detail(request, post_pk):

    currentFair = CurrentFair.objects.first()

    # Check if the user is a moderator
    is_moderator = request.user.groups.filter(name='moderator').exists()

    performance = Performance.objects.prefetch_related("instructors", "students").get(pk=post_pk)

    performance_user_organization = performance.user.organization

    other_languoid = Languoid.objects.get(name='Other')

    performance_includes_other_languoid = performance.languoids.filter(pk=other_languoid.pk).exists()

    template = 'poster_detail.html'
    context = {
        'currentFair': currentFair.name,
        'moderator': is_moderator,
        'performance': performance,
        'organization': performance_user_organization,
        'includes_other_languoid': performance_includes_other_languoid
    }
    return render(request, template, context)


@login_required
def poster_edit(request, post_pk):
    currentFair = CurrentFair.objects.first()

    performance = Performance.objects.prefetch_related("instructors", "students", "accessories").get(pk=post_pk)

    # Redirect if performance status is not 'in_progress'
    if performance.status != 'in_progress':
        return redirect('poster_detail', post_pk=post_pk)

    performance_user_organization = performance.user.organization

    owning_user = performance.user

    if request.method == "POST":
        form = PosterForm(request.POST, instance=performance)
        if form.is_valid():
            performance = form.save(commit=False)
            performance.instructors_status = "completed"
            performance.students_status = "completed"
            performance.accessories_status = "completed"
            performance.review_status = "completed"
            performance.modified_by = request.user.get_username()
            performance.save()
            return redirect("../")
    else:
        form = PosterForm(instance=performance)

    template = 'poster_edit.html'
    context = {
        'currentFair': currentFair.name,
        'performance': performance,
        'owning_user': owning_user,
        'organization': performance_user_organization,
        'form': form
    }
    return render(request, template, context)

@login_required
@user_passes_test(is_moderator)
def fair_detail(request, fair_pk=None):

    # Check if the user is a moderator
    is_moderator = request.user.groups.filter(name='moderator').exists()

    currentFair = CurrentFair.objects.first()

    if fair_pk is None:
        fair = currentFair.fair
    else:
        fair = Fair.objects.get(pk=fair_pk)

    # find the number of performances that are approved
    performances_approved = Performance.objects.filter(fair=fair).filter(status="approved")
    performances_approved_count = performances_approved.count()

    # find the number of performances that are submitted
    performances_submitted = Performance.objects.filter(fair=fair).filter(status="submitted")
    performances_submitted_count = performances_submitted.count()

    performances_total_count = performances_approved_count + performances_submitted_count

    # find the number of approved performances in each category
    performances_by_category = {}
    categories = Category.objects.filter(fair=fair)
    for category in categories:
        performances_by_category[category.name] = {
            "approved": performances_approved.filter(category=category).count(),
            "submitted": performances_submitted.filter(category=category).count(),
            "all": performances_approved.filter(category=category).count() + performances_submitted.filter(category=category).count()
        }

    # find the number of performances that are approved for each language, and store only the non-zero results
    performances_by_language = {}
    languoids = Languoid.objects.filter(fair=fair)
    for languoid in languoids:
        performances_by_language[languoid.name] = {
            "approved": performances_approved.filter(languoids=languoid).count(),
            "submitted": performances_submitted.filter(languoids=languoid).count(),
            "all": performances_approved.filter(languoids=languoid).count() + performances_submitted.filter(languoids=languoid).count()
        }
    # remove any languages with no performances
    performances_by_language = {k: v for k, v in performances_by_language.items() if v['all'] != 0}

    # find the number of performances that are approved for each grade range
    performances_by_grade_range = {}
    grade_ranges = Performance.GRADE_RANGES
    for grade_range, grade_range_display_value in grade_ranges:
        performances_by_grade_range[grade_range_display_value] = {
            "approved": performances_approved.filter(grade_range=grade_range).count(),
            "submitted": performances_submitted.filter(grade_range=grade_range).count(),
            "all": performances_approved.filter(grade_range=grade_range).count() + performances_submitted.filter(grade_range=grade_range).count()
        }
    
    # filter performances_approved and performances_submitted to only include performances with a category that is not a material submission
    performances_approved_non_material = performances_approved.filter(category__material_submission=False)
    performances_submitted_non_material = performances_submitted.filter(category__material_submission=False)

    # find the total number of tshirts needed in each size for all the approved or submitted performances
    tshirt_sizes_summary = {}
    tshirt_sizes = Student.TSHIRT_SIZES
    for tshirt_size, tshirt_size_display_value in tshirt_sizes:
        tshirt_sizes_summary[tshirt_size_display_value] = {
            "approved": performances_approved_non_material.filter(students__tshirt_size=tshirt_size).values("students").distinct().count(),
            "submitted": performances_submitted_non_material.filter(students__tshirt_size=tshirt_size).values("students").distinct().count(),
            "all": performances_approved_non_material.filter(students__tshirt_size=tshirt_size).values("students").distinct().count() + performances_submitted_non_material.filter(students__tshirt_size=tshirt_size).values("students").distinct().count()
        }

    # filter performances_approved and performances_submitted to only include performances with a category that is a material submission
    performances_approved_material = performances_approved.filter(category__material_submission=True)
    performances_submitted_material = performances_submitted.filter(category__material_submission=True)


    ## find the number of students that are in an approved performance that is a material submission, but are not in a non-material submission
    # Get all students in approved performances that are material submissions
    # students_in_approved_material_submissions = performances_approved_material.values("students").distinct()
    students_in_approved_material_submissions = Student.objects.filter(performance_student__in=performances_approved_material).distinct()

    # # print a list of students in approved material submissions
    # print('students_in_approved_material_submissions')
    # students_list = students_in_approved_material_submissions.values_list('id', 'firstname', 'lastname')
    # # Get the list of students sorted by id
    # sorted_students_list = sorted(students_list, key=lambda x: x[0])

    # # Pretty print the sorted list of students
    # for id, firstname, lastname in sorted_students_list:
    #     print(f"ID: {id}, First Name: {firstname}, Last Name: {lastname}")
    # print("length of sorted_students_list")
    # print(len(sorted_students_list))
    # print(students_in_approved_material_submissions.count())

    # Get all students in approved performances that are non-material submissions
    # students_in_approved_non_material_submissions = performances_approved_non_material.values("students").distinct()
    students_in_approved_non_material_submissions = Student.objects.filter(performance_student__in=performances_approved_non_material).distinct()


    # students_list = students_in_approved_non_material_submissions.values_list('id', 'firstname', 'lastname')
    # # Get the list of students sorted by id
    # sorted_students_list = sorted(students_list, key=lambda x: x[0])

    # # Pretty print the sorted list of students
    # for id, firstname, lastname in sorted_students_list:
    #     print(f"ID: {id}, First Name: {firstname}, Last Name: {lastname}")
    # print("length of non material sorted_students_list")
    # print(len(sorted_students_list))

    # print(students_in_approved_non_material_submissions.count())

    # Find students that are in an approved performance that is a material submission, but are not in a non-material submission
    students_in_approved_material_not_in_non_material = students_in_approved_material_submissions.exclude(id__in=students_in_approved_non_material_submissions).distinct()

    # print('students_in_approved_material_not_in_non_material')
    # print(students_in_approved_material_not_in_non_material.count())
    # print("students in both")
    # print(students_in_approved_material_submissions.filter(id__in=students_in_approved_non_material_submissions).count())

    # # find the intersection of the two sets
    # print("intersection")
    # print(students_in_approved_material_submissions.filter(id__in=students_in_approved_non_material_submissions).count())

    # # find the difference of the two sets
    # print("difference")
    # print(students_in_approved_material_submissions.difference(students_in_approved_non_material_submissions).count())


    ## find the number of students that are in a submitted performance that is a material submission, but are not in a non-material submission
    # Get all students in submitted performances that are material submissions
    students_in_submitted_material_submissions = performances_submitted_material.values("students").distinct()

    # Get all students in submitted performances that are non-material submissions
    students_in_submitted_non_material_submissions = performances_submitted_non_material.values("students").distinct()

    # Find students that are in a submitted performance that is a material submission, but are not in a non-material submission
    students_in_submitted_material_not_in_non_material = students_in_submitted_material_submissions.exclude(id__in=students_in_submitted_non_material_submissions)

    # find number of students in material submissions for bag counts
    bag_count = {
            "approved": students_in_approved_material_not_in_non_material.count(),
            "submitted": students_in_submitted_material_not_in_non_material.count(),
            "all": students_in_approved_material_not_in_non_material.count() + students_in_submitted_material_not_in_non_material.count()
        }

    # filter PerformanceAccessory instances to only include those related to performances that are approved or submitted
    performance_accessories_approved = PerformanceAccessory.objects.filter(performance__in=performances_approved)
    performance_accessories_submitted = PerformanceAccessory.objects.filter(performance__in=performances_submitted)

    # filter performance_accessories_approved and performance_accessories_submitted to only include those related to performances with a category that is not a material submission
    performance_accessories_approved_non_material = performance_accessories_approved.filter(performance__category__material_submission=False)
    performance_accessories_submitted_non_material = performance_accessories_submitted.filter(performance__category__material_submission=False)

    # find the total number of accessories in each type of accessory for all the approved performances
    accessories_summary = {}
    accessories = Accessory.objects.filter(fair=fair)
    for accessory in accessories:
        accessories_summary[accessory.name] = {
            "approved": performance_accessories_approved_non_material.filter(accessory=accessory).aggregate(Sum('count'))['count__sum'],
            "submitted": performance_accessories_submitted_non_material.filter(accessory=accessory).aggregate(Sum('count'))['count__sum'],
            "all": performance_accessories_approved_non_material.filter(accessory=accessory).aggregate(Sum('count'))['count__sum'] + performance_accessories_submitted_non_material.filter(accessory=accessory).aggregate(Sum('count'))['count__sum']
        }

    template = 'fair_detail.html'
    context = {
        'currentFair': currentFair.name,
        'fair': fair,
        'moderator': is_moderator,
        'performances_submitted_count': performances_submitted_count,
        'performances_approved_count': performances_approved_count,
        'performances_total_count': performances_total_count,
        'performances_by_category': performances_by_category,
        'performances_by_language': performances_by_language,
        'performances_by_grade_range': performances_by_grade_range,
        'tshirt_sizes_summary': tshirt_sizes_summary,
        'bag_count': bag_count,
        'accessories_summary': accessories_summary

    }
    return render(request, template, context)

# and API view that returns JSON for all the performances for the fair given by the fair_pk, with all the metadata for each performance. This is sent to the browser as a download when the user clicks the "Download All Performance data" button on the fair detail page.
class FairDownloadView(APIView):
    def get(self, request, fair_pk):
        try:
            fair = Fair.objects.get(pk=fair_pk)
            performances = Performance.objects.filter(fair=fair)

            ## Make json document
            serializer = PerformanceJsonSerializer(performances, many=True)

            # Convert the serialized data to JSON and save it to a file
            data = json.dumps(serializer.data)
            json_file_name = f'fair_{fair.name}_data.json'
            json_file = default_storage.save(json_file_name, ContentFile(data))


            ## Make xlsx for performances (non material)

            # filter performances to only include those that are approved
            performances = performances.filter(status__in=["approved"])

            non_material_submission_categories = list(Category.objects.filter(fair=fair, material_submission=False).values_list('name', flat=True))
            categories = list(Category.objects.filter(fair=fair).values_list('name', flat=True))


            performance_workbook = Workbook()

            # # Define your default font
            # default_font = Font(name='Arial')

            # # Apply the default font to all cells in the workbook
            # for sheet in performance_workbook:
            #     for row in sheet.iter_rows():
            #         for cell in row:
            #             cell.font = default_font



            for category in categories:
                # sanitize the category name to remove any characters that are not allowed in a sheet name
                category_name = re.sub(r'[\\/*?[\]:]', '_', category)
                # create a new sheet
                performance_workbook.create_sheet(title=category_name)
                # make the sheet active
                performance_workbook.active = performance_workbook[category_name]
                sheet = performance_workbook.active
                # merge the first 7 columns of the first row
                sheet.merge_cells('A1:G1')
                # set the value of the merged cells to the category name
                sheet['A1'] = category
                # set the font size of the merged cells to 20 and bold
                sheet['A1'].font = Font(size=20, bold=True)
                # set the height of the first row to 30
                sheet.row_dimensions[1].height = 30
                # set the background color of the first 7 columns of the second row to blue, and the text color to white
                for cell in sheet['A2:G2']:
                    for c in cell:
                        c.fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
                        c.font = Font(color="FFFFFF")
                # list of column headers
                headers = ["Title", "Organization", "Performance group", "Language", "Grade Range", "Performance type", "Student count"]
                # set the values of the second row to the column headers
                for i, header in enumerate(headers):
                    sheet.cell(row=2, column=i+1, value=header)
                ## add data to the sheet, starting at the third row. the data are non material performances with the current category
                # Iterate over the performances in the category
                performances_in_category = performances.filter(category__name=category)
                for performance in performances_in_category:
                    # Create a list for the current row
                    row = [
                        performance.title,
                        performance.user.organization,
                        performance.group,
                        ", ".join([languoid.name for languoid in performance.languoids.all()]),
                        performance.get_grade_range_display(),
                        performance.get_performance_type_display(),
                        performance.students.count()
                    ]
                    # Add the row to the data list
                    sheet.append(row)

                ## Adjust width of columns
                # Iterate over the columns
                for column in range(1, 8):
                    max_length = 0
                    column = get_column_letter(column)
                    for cell in sheet[column]:
                        try: 
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 1)
                    sheet.column_dimensions[column].width = adjusted_width

            # remove the default sheet
            performance_workbook.remove(performance_workbook['Sheet'])

            # Create a BytesIO object and save the workbook to it
            xlsx_file_io = BytesIO()
            performance_workbook.save(xlsx_file_io)

            # Go back to the start of the BytesIO object
            xlsx_file_io.seek(0)

            # Save the BytesIO object to a file in default storage
            xlsx_file_name = f'Fair{fair.name}-PerformanceCounts.xlsx'
            xlsx_file = default_storage.save(xlsx_file_name, ContentFile(xlsx_file_io.read()))

            # Get the full path of the files
            json_file_path = default_storage.path(json_file)
            print(json_file_path)
            xlsx_file_path = default_storage.path(xlsx_file)
            print(xlsx_file_path)

            zip_folder_name = f'fair_{fair.name}_data/'

            # Create a new zip file
            zip_file_name = f'fair_{fair.name}_data.zip'
            with zipfile.ZipFile(zip_file_name, 'w') as zip_file:
                # Add the json file to the zip file
                zip_file.write(json_file_path, arcname=zip_folder_name+json_file_name)

                # Add the xlsx file to the zip file
                zip_file.write(xlsx_file_path, arcname=zip_folder_name+xlsx_file_name)

            # Create a generator that reads the file and yields the data
            def file_iterator():
                try:
                    with open(zip_file_name, 'rb') as f:
                        for chunk in iter(lambda: f.read(4096), b''):
                            yield chunk
                finally:
                    os.remove(zip_file_name)

            # Create a StreamingHttpResponse that uses the file_iterator
            response = StreamingHttpResponse(file_iterator(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{zip_file_name}"'

            # Delete the files from the default storage
            default_storage.delete(json_file)
            default_storage.delete(xlsx_file)

            return response

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# and API view that returns JSON for all the performances for the fair given by the fair_pk, with all the metadata for each performance. This is sent to the browser as a download when the user clicks the "Download All Performance data" button on the fair detail page.
class JudgeSheetsDownloadView(APIView):
    def get(self, request, fair_pk):
        fair = Fair.objects.get(pk=fair_pk)
        performances = Performance.objects.filter(fair=fair)

        ## Make xlsx for performances (non material)

        # filter performances to only include those that are approved
        performances = performances.filter(status__in=["approved"])

        # filter performances by category, exluding the categories "Poster", "Comics and Cartoons", "Mobile Video"
        performances = performances.exclude(category__name__in=["Poster", "Comics and Cartoons", "Mobile Video"])

        # sort performances by category, then by organization, then by grade range, then by group, then by title
        performances = performances.order_by('category__name', 'user__organization', 'grade_range', 'group', 'title')

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="fair_{fair.name}_judging_sheets.pdf"'
        
        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter

        # Register the DejaVuSans font
        font_path = os.path.join(settings.STATIC_ROOT, 'DejaVuSans.ttf')
        # font_bold_path = os.path.join(settings.STATIC_ROOT, 'DejaVuSans-Bold.ttf')
        pdfmetrics.registerFont(TTFont('DejaVuSans', font_path))
        # pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', font_bold_path))

        def draw_static_elements():
            # Title banner image
            # Get image
            image_path = os.path.join(settings.STATIC_ROOT, 'onaylf.png')
            # Create an Image object
            image = Image(image_path)
            # Draw the image
            image.drawOn(p, 30, height-100)
            # p.drawString(30, height-30, "SAM NOBLE MUSEUM DEPARTMENT OF NATIVE AMERICAN LANGUAGES")
            # Subtitle with Blue Background
            sky_blue = Color(0.429, 0.708, 0.982)  # RGB values for light blue
            p.setFillColor(sky_blue)
            rectangle_width = width * 0.8  # 80% of the page width
            rectangle_x = (width - rectangle_width) / 2  # Calculate the x-coordinate to center the rectangle

            p.rect(rectangle_x, height-140, rectangle_width, 30, fill=True, stroke=False)
            p.setFillColor(white)
            p.setFont("Helvetica", 14)
            p.drawString(130, height-130, "Oklahoma Native American Youth Language Fair 2024")

            p.setFillColor(black)
            # Horizontal Rule after Dynamic Text
            p.line(30, height-300, width-30, height-300)

            p.setFont("Helvetica", 9)
            p.drawString(35, height-315, f"Please rate the {performance.category.name}")
            p.drawString(35, height-325, "using a scale of 1 to 5.")
            p.drawString(435, height-315, "Circle score for each")
            p.drawString(435, height-325, "critera:")
            p.setFont("Helvetica-Bold", 9)
            p.drawString(35, height-345, "1 = Poor")
            p.drawString(35, height-355, "2 = Below Average")
            p.drawString(35, height-365, "3 = Average")
            p.drawString(35, height-375, "4 = Above Average")
            p.drawString(35, height-385, "5 = Excellent")

            # Likert-style Ratings
            ratings = ["                                                   Use of Language:",
                       "Originality & Creativity in Illustration & Content:",
                       "                                                                       Effort:"]
            y_position = height - 345  # Subtract 100 from the y-coordinate
            for rating in ratings:
                p.drawString(220, y_position, rating)
                for i in range(1, 6):
                    p.drawString(420 + 15*i, y_position, str(i))
                y_position -= 20

            # Horizontal Rule after Likerts
            p.line(30, y_position-10, width-30, y_position-10)

            # Total Score Box, Comments Box, Judge Name and Signature
            p.drawString(350, y_position-20, "Total Score:")
            p.rect(420, y_position-40, 100, 30)
            p.drawString(50, y_position-60, "Comments:")
            p.rect(48, y_position-225, width-100, 150)
            p.drawString(130, y_position-280, "Judge Name:")
            p.line(200, y_position-280, width-130, y_position-280)
            p.drawString(130, y_position-320, "Signature:")
            p.line(200, y_position-320, width-130, y_position-320)
            

            # Horizontal Rule before footer
            p.line(30, 40, width-30, 40)
            # Footer
            p.setFont("Helvetica", 8)
            p.drawString(30, 30, f"Oklahoma Native American Youth Language Fair {fair.name} Judging Sheet - {performance.category.name}")

        for performance in performances:
            logger.info(performance.id)
            logger.info(performance.title)
            draw_static_elements()

            # Dynamic text drawing goes here
            p.setFont("DejaVuSans", 10)
            p.drawString(30, height-155, "Program/School: ")
            p.drawString(150, height-155, f"{performance.user.organization}")
            p.drawString(30, height-175, "Grade: ")
            p.drawString(150, height-175, f"{performance.get_grade_range_display()}")
            p.drawString(30, height-195, "Presenting Group: ")
            p.drawString(150, height-195, f"{performance.group}")
            p.drawString(30, height-215, "Title: ")
            p.drawString(150, height-215, f"{performance.title}")
            p.drawString(30, height-235, "Language: ")
            p.drawString(150, height-235, f"{', '.join([languoid.name for languoid in performance.languoids.all()])}")
            p.drawString(30, height-255, "Category: ")
            p.drawString(150, height-255, f"{performance.category.name}")
            p.drawString(30, height-275, "Type: ")
            p.drawString(150, height-275, f"{performance.get_performance_type_display()}")

            p.showPage()

        p.save()
        logger.info("server finished")
        return response


# def query_inveniordm(request):
#     # The URL to the InvenioRDM records API endpoint
#     url = "https://143.244.215.98/api/records"

#     # Parameters for the search query
#     params = {
#         "q": "metadata.languages.props.family:(\"coch1271\")"
#     } 

# # https://143.244.215.98/api/records?q=%22rood%22%20AND%20metadata.languages.id%3A%28%22wich1260%22%29


#     # Make a GET request to the API
#     try:
#         response = requests.get(url, params=params, verify=False)  # Use verify=False only if necessary for self-signed certificates
#     except requests.exceptions.RequestException as e:
#         # Handle request exceptions (e.g., network errors)
#         return JsonResponse({"error": "Network error or invalid URL", "details": str(e)}, status=500)

#     # Check if the request was successful
#     if response.status_code == 200:
#         # Parse the response JSON and return it
#         data = response.json()
#         return JsonResponse(data, safe=False)
#     else:
#         # Handle errors or unsuccessful responses
#         error_message = response.text
#         try:
#             error_data = response.json()
#             error_message = error_data.get('message', response.text)
#         except ValueError:
#             pass  # JSON parsing failed, use the raw text
#         return JsonResponse({"error": "Failed to fetch data", "details": error_message}, status=response.status_code)






# performances_all = Performance.objects.filter(status__in=["approved", "submitted"])

# performances_non = performances_all.filter(Q(category__name="Master Performer") | Q(category__name="Modern Song") | Q(category__name="Skit/Short Play") | Q(category__name="Spoken Language") | Q(category__name="Spoken Poetry") | Q(category__name="Spoken Prayer") | Q(category__name="Traditional Song"))

# students_non = Student.objects.filter(performance_student__in=performances_non).distinct()


# performances_material = performances_all.filter(Q(category__name="Books") | Q(category__name="Comics and Cartoons") | Q(category__name="Film and Video") | Q(category__name="Mobile Video") | Q(category__name="Poster") | Q(category__name="Puppet Show"))

# students_material = Student.objects.filter(performance_student__in=performances_material).distinct()

# # find the students that are in both performances and performances_material
# students_both = students_non.filter(id__in=students_material)
